# backend/journals/bulk_import_template.py

"""
Generates a downloadable Excel template for bulk journal entry import.

The template has 3 sheets:
    1. Instructions   — Rules, column reference, examples
    2. Journal Entries — User fills this in (the actual data sheet)
    3. Account Reference — Auto-generated tree of the company's CoA

CALLED FROM:
    journals/views.py → BulkImportTemplateDownloadView
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from chartofaccounts.models import AccountClassification, Account


# ── Style constants ──
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
REQUIRED_FILL = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
OPTIONAL_FILL = PatternFill(start_color='555555', end_color='555555', fill_type='solid')
EXAMPLE_FILL = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
REF_HEADER_FILL = PatternFill(start_color='264653', end_color='264653', fill_type='solid')
TREE_L1_FILL = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
TREE_L2_FILL = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1A1A2E')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=11, color='333333')
NOTE_FONT = Font(name='Arial', italic=True, size=9, color='666666')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def generate_bulk_import_template(company):
    wb = openpyxl.Workbook()
    _build_instructions_sheet(wb, company)
    _build_journal_entries_sheet(wb)
    _build_account_reference_sheet(wb, company)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_instructions_sheet(wb, company):
    ws = wb.create_sheet('Instructions', 0)
    ws.sheet_properties.tabColor = '1A1A2E'
    ws.column_dimensions['A'].width = 90
    row = 1
    ws.cell(row=row, column=1, value='Bulk Journal Entry Import — Instructions').font = TITLE_FONT
    row += 2
    ws.cell(row=row, column=1, value=f'Company: {company.name}').font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f'Base Currency: {company.base_currency}').font = BODY_FONT
    row += 2
    ws.cell(row=row, column=1, value='GENERAL RULES').font = SUBTITLE_FONT
    row += 1

    rules = [
        '1. Fill in the "Journal Entries" sheet (Sheet 2). Do not modify column headers.',
        '2. Each journal entry is identified by its Entry Prefix + Entry Number combination.',
        '3. Rows sharing the same prefix + number form one journal entry.',
        '4. Header fields (Date, Description, etc.) only need to be filled on the FIRST row of each entry.',
        '5. Subsequent rows of the same entry only need Account Name, Debit, and Credit.',
        '6. Each entry must have at least 2 lines.',
        '7. Each entry must be balanced: total Debit must equal total Credit.',
        '8. All entries will be created as DRAFT. You can post them from the system after review.',
        '',
        'ENTRY PREFIX (Optional)',
        '   The prefix for the entry number (e.g., "JE-", "MJ-", "OB-").',
        '   Can be left blank. If blank, the entry number is just the number part.',
        '   Example: prefix "JE-" + number "0050" = entry number "JE-0050".',
        '   Example: blank prefix + number "0050" = entry number "0050".',
        '',
        'ENTRY NUMBER RULES',
        '   Entry Number is required on every row.',
        '   You may use any prefix or leave it blank.',
        '   The combined entry number (prefix + number) must not match any existing journal entry.',
        '   The combined entry number must not appear in separate, non-contiguous blocks in the file.',
        '   If an entry number appears in non-contiguous sections, ALL occurrences are rejected.',
        '',
        'ACCOUNT NAME RULES',
        '   Account names must match exactly (case-sensitive) with existing accounts.',
        '   Refer to the "Account Reference" sheet (Sheet 3) for the list of available accounts.',
        '   Only active accounts can receive journal entries.',
        '',
        'DEBIT / CREDIT RULES',
        '   Each row must have EITHER a Debit amount OR a Credit amount — not both.',
        '   Leave the other column empty or put zero — zero is treated as empty.',
        '   Amounts must be positive numbers greater than zero.',
        '   Maximum amount per line: 999,999,999,999.99.',
        '',
        'DATE FORMAT',
        '   Preferred format: YYYY-MM-DD (e.g., 2026-03-10)',
        '   Also accepted: DD/MM/YYYY, MM/DD/YYYY, YYYY/MM/DD, DD-MM-YYYY',
        '   Excel date cells are also accepted automatically.',
        '',
        'CURRENCY & EXCHANGE RATE (Optional)',
        f'   If left blank, the company base currency ({company.base_currency}) is used with rate 1.0.',
        '   If you specify a currency other than the base, you must also provide an exchange rate.',
        '   Exchange rate = how many units of base currency per 1 unit of the specified currency.',
        f'   Example: If base is {company.base_currency} and currency is USD with rate 120, '
        f'it means 1 USD = 120 {company.base_currency}.',
        '',
        'JOURNAL TYPE (Optional, NOT case-sensitive)',
        '   Valid values: ADJUSTMENT, PURCHASE, SALES, PAYROLL, DEPRECIATION, INVESTMENT,',
        '   DIVIDEND, TAX, OPENING_BALANCE, TRANSFER, CURRENCY_EXCHANGE, OTHER',
        '   You can type "payroll" or "Payroll" or "PAYROLL" — all work the same.',
        '   Leave blank if not applicable.',
        '',
        'FILE FORMAT',
        '   You can upload this .xlsx file or a .csv file with the same column structure.',
        '   For CSV: use comma delimiter and UTF-8 encoding.',
    ]

    for rule in rules:
        ws.cell(row=row, column=1, value=rule).font = BODY_FONT if rule else NOTE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='COLUMN REFERENCE').font = SUBTITLE_FONT
    row += 1

    columns = [
        ('Entry Prefix (Optional)', 'Prefix for entry number. e.g., "JE-", "MJ-". Can be blank.'),
        ('Entry Number (Required)', 'Number part. e.g., "0050". Combined with prefix = "JE-0050". Required on every row.'),
        ('Date (Required)', 'Transaction date. See DATE FORMAT above. Only on first row of each entry.'),
        ('Description (Required)', 'What this journal entry is for. Only on first row of each entry.'),
        ('Reference (Optional)', 'External reference number. Only on first row.'),
        ('Journal Type (Optional)', 'Category. Not case-sensitive. Only on first row.'),
        ('Currency (Optional)', 'ISO 4217 code (e.g., USD, EUR). Defaults to base currency.'),
        ('Exchange Rate (Optional)', 'Rate to base currency. Defaults to 1.0.'),
        ('Account Name (Required)', 'Must match an existing active account name exactly (case-sensitive).'),
        ('Debit (Required*)', 'Debit amount. Leave blank or zero if this line is a credit.'),
        ('Credit (Required*)', 'Credit amount. Leave blank or zero if this line is a debit.'),
        ('Line Description (Optional)', 'Note for this specific line.'),
    ]

    for col_name, col_desc in columns:
        ws.cell(row=row, column=1, value=f'  {col_name}: {col_desc}').font = BODY_FONT
        row += 1

    ws.cell(row=row + 1, column=1,
            value='* Each row requires exactly one of Debit or Credit to have a positive value. '
                  'Zero is treated as empty.').font = NOTE_FONT


def _build_journal_entries_sheet(wb):
    ws = wb.create_sheet('Journal Entries', 1)
    ws.sheet_properties.tabColor = '2D6A4F'

    headers = [
        ('Entry Prefix (Optional)', 14),
        ('Entry Number (Required)', 16),
        ('Date (Required)', 14),
        ('Description (Required)', 35),
        ('Reference (Optional)', 18),
        ('Journal Type (Optional)', 18),
        ('Currency (Optional)', 12),
        ('Exchange Rate (Optional)', 16),
        ('Account Name (Required)', 35),
        ('Debit (Required*)', 16),
        ('Credit (Required*)', 16),
        ('Line Description (Optional)', 30),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = REQUIRED_FILL if '(Required' in header else OPTIONAL_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    examples = [
        ['JE-', '0050', '2026-03-10', 'Office rent for March 2026', 'RENT-MAR',
         'ADJUSTMENT', '', '', 'Rent Expense', 25000.00, '', 'March rent'],
        ['JE-', '0050', '', '', '', '', '', '', 'Bank', '', 25000.00, 'Paid from bank'],
        ['', '0051', '2026-03-11', 'Salary payment', '', 'payroll',
         '', '', 'Salaries and Wages', 15000.00, '', 'Sales department'],
        ['', '0051', '', '', '', '', '', '', 'Salaries and Wages', 10000.00, '', 'Admin department'],
        ['', '0051', '', '', '', '', '', '', 'Bank', '', 25000.00, 'From bank account'],
    ]

    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
            cell.font = BODY_FONT
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    ws.cell(row=8, column=1,
            value='↑ Example entries above. Entry 1 has prefix "JE-", Entry 2 has no prefix. '
                  'Replace with your data.').font = NOTE_FONT
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'


def _build_account_reference_sheet(wb, company):
    ws = wb.create_sheet('Account Reference', 2)
    ws.sheet_properties.tabColor = '264653'

    ref_headers = [
        ('Layer', 8), ('Classification / Account', 45), ('Account Code', 14),
        ('Normal Balance', 14), ('Currency', 10), ('Type', 12),
    ]

    for col_idx, (header, width) in enumerate(ref_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 2
    classifications = AccountClassification.objects.filter(company=company).order_by('internal_path')
    accounts = Account.objects.filter(company=company, is_active=True).select_related('classification').order_by('internal_path')

    accts_by_class = {}
    for acct in accounts:
        accts_by_class.setdefault(acct.classification_id, []).append(acct)

    for classification in classifications:
        layer = classification.layer
        indent = '    ' * (layer - 1)
        ws.cell(row=row, column=1, value=f'L{layer}').font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=f'{indent}{classification.name}').font = BOLD_FONT
        ws.cell(row=row, column=6, value='Classification').font = NOTE_FONT

        if layer == 1:
            for col in range(1, 7): ws.cell(row=row, column=col).fill = TREE_L1_FILL
        elif layer == 2:
            for col in range(1, 7): ws.cell(row=row, column=col).fill = TREE_L2_FILL
        for col in range(1, 7): ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        if layer == 3 and classification.id in accts_by_class:
            for acct in accts_by_class[classification.id]:
                path_parts = acct.internal_path.split('.')
                acct_depth = len(path_parts) - 3
                acct_indent = '    ' * 3 + '  ' * acct_depth
                prefix = '└─ ' if acct.is_sub_account else ''
                ws.cell(row=row, column=1, value=f'L{3 + acct_depth + 1}').font = Font(name='Arial', size=9, color='888888')
                ws.cell(row=row, column=2, value=f'{acct_indent}{prefix}{acct.name}').font = BODY_FONT
                ws.cell(row=row, column=3, value=acct.code).font = BODY_FONT
                ws.cell(row=row, column=4, value=acct.normal_balance).font = BODY_FONT
                ws.cell(row=row, column=5, value=acct.currency).font = BODY_FONT
                acct_type = 'System' if acct.is_system_account else 'Custom'
                if acct.is_sub_account: acct_type += ' (Sub)'
                ws.cell(row=row, column=6, value=acct_type).font = NOTE_FONT
                for col in range(1, 7): ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1

    ws.freeze_panes = 'A2'
    ws.cell(row=row + 1, column=2,
            value="This sheet is auto-generated from your company's Chart of Accounts. Do not edit.").font = NOTE_FONT