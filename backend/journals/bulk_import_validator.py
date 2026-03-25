# backend/journals/bulk_import_validator.py

"""
Validates an uploaded bulk journal import file (Excel or CSV).

THE VALIDATION PIPELINE:
    Phase 0 — File Format: Can we parse it? Is it .xlsx or .csv?
    Phase 1 — Header Check: Do the column headers match?
    Phase 2 — Row Parsing & Grouping: Group rows by entry number.
    Phase 3 — Per-Group Validation: Balance, accounts, amounts, dates.
    Phase 4 — Cross-Group Validation: Duplicate numbers, existing DB conflicts.

RETURN VALUE:
    {
        "valid": True/False,
        "summary": {"total_groups": N, "accepted": N, "rejected": N},
        "accepted_entries": [...],
        "rejected_entries": [...],
        "parsed_data": [...]  (only for accepted entries, used by service)
    }
"""

import csv
import io
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl

from chartofaccounts.models import Account
from journals.models import ManualJournal, JournalTypeChoices


# ── Expected column headers ──
# Entry Prefix changed from Required to Optional
EXPECTED_HEADERS = [
    'Entry Prefix (Optional)',
    'Entry Number (Required)',
    'Date (Required)',
    'Description (Required)',
    'Reference (Optional)',
    'Journal Type (Optional)',
    'Currency (Optional)',
    'Exchange Rate (Optional)',
    'Account Name (Required)',
    'Debit (Required*)',
    'Credit (Required*)',
    'Line Description (Optional)',
]

# Column indices (0-based)
COL_PREFIX = 0
COL_NUMBER = 1
COL_DATE = 2
COL_DESCRIPTION = 3
COL_REFERENCE = 4
COL_JOURNAL_TYPE = 5
COL_CURRENCY = 6
COL_EXCHANGE_RATE = 7
COL_ACCOUNT_NAME = 8
COL_DEBIT = 9
COL_CREDIT = 10
COL_LINE_DESC = 11

VALID_JOURNAL_TYPES = {choice[0] for choice in JournalTypeChoices.choices}


# ══════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════

def validate_bulk_import(file_obj, file_name, company):
    """
    Validate the uploaded file and return results.
    """
    # ── Phase 0: Parse file ──
    rows, parse_error = _parse_file(file_obj, file_name)
    if parse_error:
        return {
            'valid': False,
            'file_error': parse_error,
            'summary': {'total_groups': 0, 'accepted': 0, 'rejected': 0},
            'accepted_entries': [],
            'rejected_entries': [],
            'parsed_data': [],
        }

    # ── Phase 1: Header check ──
    header_error = _validate_headers(rows[0] if rows else [])
    if header_error:
        return {
            'valid': False,
            'file_error': header_error,
            'summary': {'total_groups': 0, 'accepted': 0, 'rejected': 0},
            'accepted_entries': [],
            'rejected_entries': [],
            'parsed_data': [],
        }

    # ── Phase 2: Group rows by entry number ──
    data_rows = rows[1:]  # Skip header
    groups, grouping_errors = _group_rows(data_rows)

    # ── Phase 3 & 4: Validate groups + cross-group checks ──
    result = _validate_groups(groups, company, grouping_errors, data_rows)

    return result


# ══════════════════════════════════════════════════
# PHASE 0: FILE PARSING
# ══════════════════════════════════════════════════

def _parse_file(file_obj, file_name):
    ext = file_name.lower().rsplit('.', 1)[-1] if '.' in file_name else ''
    if ext == 'xlsx':
        return _parse_xlsx(file_obj)
    elif ext == 'csv':
        return _parse_csv(file_obj)
    else:
        return [], (
            f'Unsupported file format: ".{ext}". '
            f'Please upload an .xlsx or .csv file.'
        )


def _parse_xlsx(file_obj):
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception:
        return [], (
            'Cannot read the uploaded file. '
            'Please ensure it is a valid .xlsx file.'
        )

    # Find the Journal Entries sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower().replace(' ', '') == 'journalentries':
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=1, max_col=12, values_only=True):
        cleaned = []
        for val in row:
            if val is None:
                cleaned.append(None)
            else:
                s = str(val).strip()
                cleaned.append(s if s else None)
        rows.append(cleaned)

    # Remove trailing empty rows
    while rows and all(v is None for v in rows[-1]):
        rows.pop()

    if len(rows) < 2:
        return [], 'The file has no data rows. Please add journal entries below the header row.'

    return rows, None


def _parse_csv(file_obj):
    try:
        if hasattr(file_obj, 'read'):
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8-sig')
        else:
            content = str(file_obj)

        reader = csv.reader(io.StringIO(content))
        rows = []
        for row in reader:
            cleaned = []
            for val in row:
                s = val.strip() if val else None
                cleaned.append(s if s else None)
            rows.append(cleaned)

        # Pad short rows to 12 columns
        for i, row in enumerate(rows):
            while len(row) < 12:
                row.append(None)

        while rows and all(v is None for v in rows[-1]):
            rows.pop()

        if len(rows) < 2:
            return [], 'The CSV file has no data rows.'

        return rows, None
    except Exception as e:
        return [], f'Cannot parse the CSV file: {str(e)}'


# ══════════════════════════════════════════════════
# PHASE 1: HEADER VALIDATION
# ══════════════════════════════════════════════════

def _validate_headers(header_row):
    if not header_row or len(header_row) < 12:
        return (
            'Missing or incomplete header row. '
            'The file must have 12 columns matching the template. '
            'Please download the template and use it as the base.'
        )

    for i, expected in enumerate(EXPECTED_HEADERS):
        actual = header_row[i] if i < len(header_row) and header_row[i] else ''
        if actual != expected:
            return (
                f'Column {i + 1} header mismatch. '
                f'Expected: "{expected}", found: "{actual}". '
                f'Do not modify the header row. Download a fresh template if needed.'
            )

    return None


# ══════════════════════════════════════════════════
# PHASE 2: ROW GROUPING
# ══════════════════════════════════════════════════

def _group_rows(data_rows):
    """
    Group rows by Entry Prefix + Entry Number.
    Entry Prefix is OPTIONAL — can be blank.
    Entry Number is REQUIRED on every data row.
    """
    groups = {}  # entry_key → list of (excel_row_number, row_data)
    errors = []

    for idx, row in enumerate(data_rows):
        excel_row = idx + 2  # 1-based, skip header

        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        # Entry prefix is optional, default to empty string
        prefix = row[COL_PREFIX] or ''
        number = row[COL_NUMBER]

        # Entry Number is required
        if not number:
            # Only report error if the row has other data
            if any(v is not None for v in row):
                if prefix:
                    errors.append({
                        'row': excel_row,
                        'column': 'Entry Number',
                        'message': (
                            f'Entry Number is required. '
                            f'Found prefix "{prefix}" but no entry number.'
                        ),
                    })
                else:
                    errors.append({
                        'row': excel_row,
                        'column': 'Entry Number',
                        'message': 'Entry Number is required on every row.',
                    })
            continue

        entry_key = f"{prefix}{number}"

        if entry_key not in groups:
            groups[entry_key] = []

        groups[entry_key].append((excel_row, row))

    return groups, errors


# ══════════════════════════════════════════════════
# PHASE 3 & 4: GROUP VALIDATION
# ══════════════════════════════════════════════════

def _validate_groups(groups, company, grouping_errors, data_rows):
    """
    Validate each group (journal entry) and perform cross-group checks.
    """
    # ── Pre-load company data ──
    account_map = {}
    for acct in Account.objects.filter(company=company, is_active=True):
        account_map[acct.name] = acct

    inactive_names = set(
        Account.objects.filter(
            company=company, is_active=False,
        ).values_list('name', flat=True)
    )

    # ── Check for existing entry numbers in the database ──
    all_entry_keys = set(groups.keys())
    existing_numbers = set(
        ManualJournal.objects.filter(
            company=company,
            entry_number__in=all_entry_keys,
        ).values_list('entry_number', flat=True)
    )

    # ── Detect non-contiguous duplicate entry numbers ──
    # Track the order entry keys appear row-by-row.
    # If the same key appears, disappears, then reappears, it's non-contiguous.
    seen_keys_order = []
    last_key = None
    for idx, row in enumerate(data_rows):
        if all(v is None for v in row):
            continue
        prefix = row[COL_PREFIX] or ''
        number = row[COL_NUMBER]
        if not number:
            continue
        entry_key = f"{prefix}{number}"
        if entry_key != last_key:
            seen_keys_order.append(entry_key)
            last_key = entry_key

    key_appearances = Counter(seen_keys_order)
    non_contiguous_keys = {k for k, count in key_appearances.items() if count > 1}

    accepted = []
    rejected = []

    # ── Reject non-contiguous duplicates first ──
    for bad_key in non_contiguous_keys:
        if bad_key in groups:
            row_list = groups[bad_key]
            # Build separate row ranges for each contiguous block
            blocks = []
            current_block_start = None
            current_block_end = None
            prev_row = None
            for row_num, _ in row_list:
                if prev_row is None or row_num != prev_row + 1:
                    # Start a new block
                    if current_block_start is not None:
                        blocks.append(f"{current_block_start}-{current_block_end}")
                    current_block_start = row_num
                current_block_end = row_num
                prev_row = row_num
            if current_block_start is not None:
                blocks.append(f"{current_block_start}-{current_block_end}")

            rejected.append({
                'entry_number': bad_key,
                'row_range': ', '.join(blocks),  # e.g., "2-3, 6-7"
                'errors': [{
                    'row': row_list[0][0],
                    'column': 'Entry Number',
                    'message': (
                        f'Entry number "{bad_key}" appears in multiple non-contiguous '
                        f'blocks (rows {", ".join(blocks)}) in the file. This looks like '
                        f'duplicate journal entries with the same number. All occurrences '
                        f'are rejected. Use a unique entry number for each journal.'
                    ),
                }],
            })

    # ── Validate each remaining group ──
    for entry_key, row_list in groups.items():
        # Skip non-contiguous duplicates (already rejected above)
        if entry_key in non_contiguous_keys:
            continue

        entry_errors = []

        # Check if entry number exists in the database
        if entry_key in existing_numbers:
            entry_errors.append({
                'row': row_list[0][0],
                'column': 'Entry Number',
                'message': (
                    f'Entry number "{entry_key}" already exists in the system. '
                    f'Use a different entry number.'
                ),
            })

        # Validate the group
        group_result = _validate_single_group(
            entry_key, row_list, account_map, inactive_names, company,
        )
        entry_errors.extend(group_result['errors'])

        if entry_errors:
            rejected.append({
                'entry_number': entry_key,
                'row_range': f"{row_list[0][0]}-{row_list[-1][0]}",
                'errors': entry_errors,
            })
        else:
            accepted.append(group_result['parsed'])

    # ── Add grouping-level errors ──
    if grouping_errors:
        rejected.insert(0, {
            'entry_number': '(ungrouped rows)',
            'row_range': '',
            'errors': grouping_errors,
        })

    total = len(groups)
    accepted_count = len(accepted)
    rejected_count = total - accepted_count

    return {
        'valid': rejected_count == 0 and len(grouping_errors) == 0,
        'summary': {
            'total_groups': total,
            'accepted': accepted_count,
            'rejected': rejected_count,
            'total_rows': sum(len(g) for g in groups.values()),
        },
        'accepted_entries': [
            {
                'entry_number': p['entry_number'],
                'date': str(p['date']),
                'description': p['description'],
                'lines': len(p['lines']),
            }
            for p in accepted
        ],
        'rejected_entries': rejected,
        'parsed_data': accepted,
    }


def _validate_single_group(entry_key, row_list, account_map, inactive_names, company):
    """Validate a single journal entry (one group of rows)."""
    errors = []

    # ── Minimum 2 lines ──
    if len(row_list) < 2:
        errors.append({
            'row': row_list[0][0],
            'column': 'Entry',
            'message': (
                f'Entry "{entry_key}" has only {len(row_list)} line. '
                f'A journal entry requires at least 2 lines.'
            ),
        })

    # ── Parse header fields from ALL rows in the group ──
    # Header fields can appear on ANY row. If the same field appears
    # on multiple rows with the SAME value, it's accepted. If values
    # CONFLICT across rows, the entry is rejected.

    first_row_num = row_list[0][0]

    # Collect all non-empty values for each header field.
    # Using dict to track {normalized_value: first_row_where_seen}.
    dates_found = {}
    descriptions_found = {}
    references_found = {}
    journal_types_found = {}
    currencies_found = {}
    exchange_rates_found = {}

    for row_num, row in row_list:
        if row[COL_DATE]:
            # Normalize: parse the date first, then use the date object as key
            # This ensures "2026-03-10" and "10/03/2026" are treated as same value
            try:
                parsed = _parse_date(row[COL_DATE])
                key = str(parsed)  # "2026-03-10"
            except ValueError:
                key = str(row[COL_DATE]).strip()  # Unparseable — keep raw
            if key not in dates_found:
                dates_found[key] = row_num

        if row[COL_DESCRIPTION]:
            val = str(row[COL_DESCRIPTION]).strip()
            if val not in descriptions_found:
                descriptions_found[val] = row_num

        if row[COL_REFERENCE]:
            val = str(row[COL_REFERENCE]).strip()
            if val not in references_found:
                references_found[val] = row_num

        if row[COL_JOURNAL_TYPE]:
            val = str(row[COL_JOURNAL_TYPE]).strip().upper()
            if val not in journal_types_found:
                journal_types_found[val] = row_num

        if row[COL_CURRENCY]:
            val = str(row[COL_CURRENCY]).strip().upper()
            if val not in currencies_found:
                currencies_found[val] = row_num

        if row[COL_EXCHANGE_RATE]:
            val = str(row[COL_EXCHANGE_RATE]).strip()
            if val not in exchange_rates_found:
                exchange_rates_found[val] = row_num

    # ── Date (required) ──
    parsed_date = None
    if len(dates_found) == 0:
        errors.append({
            'row': first_row_num,
            'column': 'Date',
            'message': (
                f'Date is required for entry "{entry_key}". '
                f'Fill it on at least one row of this entry.'
            ),
        })
    elif len(dates_found) > 1:
        conflict_details = [f'row {r}: "{v}"' for v, r in dates_found.items()]
        errors.append({
            'row': first_row_num,
            'column': 'Date',
            'message': (
                f'Entry "{entry_key}" has conflicting dates: '
                f'{", ".join(conflict_details)}. '
                f'All rows in an entry must have the same date '
                f'(or leave it blank on the other rows).'
            ),
        })
    else:
        date_key = list(dates_found.keys())[0]
        date_row = list(dates_found.values())[0]
        try:
            parsed_date = _parse_date(date_key)
        except ValueError:
            errors.append({
                'row': date_row,
                'column': 'Date',
                'message': (
                    f'Invalid date: "{date_key}". '
                    f'Accepted formats: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, '
                    f'YYYY/MM/DD, DD-MM-YYYY.'
                ),
            })

    # Lock date check
    if parsed_date and company.lock_date and parsed_date <= company.lock_date:
        errors.append({
            'row': first_row_num,
            'column': 'Date',
            'message': (
                f'Date {parsed_date} is on or before the lock date ({company.lock_date}). '
                f'Transactions before the lock date are frozen. '
                f'Choose a date after {company.lock_date}.'
            ),
        })

    # ── Description (required) ──
    description = None
    if len(descriptions_found) == 0:
        errors.append({
            'row': first_row_num,
            'column': 'Description',
            'message': (
                f'Description is required for entry "{entry_key}". '
                f'Fill it on at least one row of this entry.'
            ),
        })
    elif len(descriptions_found) > 1:
        conflict_details = [f'row {r}: "{v[:40]}"' for v, r in descriptions_found.items()]
        errors.append({
            'row': first_row_num,
            'column': 'Description',
            'message': (
                f'Entry "{entry_key}" has conflicting descriptions: '
                f'{", ".join(conflict_details)}. '
                f'All rows in an entry must have the same description '
                f'(or leave it blank on the other rows).'
            ),
        })
    else:
        description = list(descriptions_found.keys())[0]

    # ── Reference (optional) ──
    reference = None
    if len(references_found) > 1:
        conflict_details = [f'row {r}' for r in references_found.values()]
        errors.append({
            'row': first_row_num,
            'column': 'Reference',
            'message': (
                f'Entry "{entry_key}" has conflicting references on rows '
                f'{", ".join(conflict_details)}. '
                f'Use the same reference or fill it on only one row.'
            ),
        })
    elif len(references_found) == 1:
        reference = list(references_found.keys())[0]

    # ── Journal Type (optional, NOT case-sensitive) ──
    journal_type = None
    if len(journal_types_found) > 1:
        conflict_details = [f'row {r}: "{v}"' for v, r in journal_types_found.items()]
        errors.append({
            'row': first_row_num,
            'column': 'Journal Type',
            'message': (
                f'Entry "{entry_key}" has conflicting journal types: '
                f'{", ".join(conflict_details)}. '
                f'Use the same type or fill it on only one row.'
            ),
        })
    elif len(journal_types_found) == 1:
        jt_value = list(journal_types_found.keys())[0]
        jt_row = list(journal_types_found.values())[0]
        if jt_value not in VALID_JOURNAL_TYPES:
            errors.append({
                'row': jt_row,
                'column': 'Journal Type',
                'message': (
                    f'Invalid journal type: "{jt_value}". '
                    f'Valid values: {", ".join(sorted(VALID_JOURNAL_TYPES))}. '
                    f'Journal type is not case-sensitive.'
                ),
            })
        else:
            journal_type = jt_value

    # ── Currency (optional) ──
    currency = company.base_currency
    if len(currencies_found) > 1:
        conflict_details = [f'row {r}: "{v}"' for v, r in currencies_found.items()]
        errors.append({
            'row': first_row_num,
            'column': 'Currency',
            'message': (
                f'Entry "{entry_key}" has conflicting currencies: '
                f'{", ".join(conflict_details)}. '
                f'Use the same currency or fill it on only one row.'
            ),
        })
    elif len(currencies_found) == 1:
        curr_value = list(currencies_found.keys())[0]
        curr_row = list(currencies_found.values())[0]
        if len(curr_value) != 3:
            errors.append({
                'row': curr_row,
                'column': 'Currency',
                'message': (
                    f'Currency must be a 3-letter ISO code (e.g., USD, EUR). '
                    f'Found: "{curr_value}".'
                ),
            })
        else:
            currency = curr_value

    # ── Exchange Rate (optional) ──
    exchange_rate = Decimal('1.000000')
    if len(exchange_rates_found) > 1:
        conflict_details = [f'row {r}: {v}' for v, r in exchange_rates_found.items()]
        errors.append({
            'row': first_row_num,
            'column': 'Exchange Rate',
            'message': (
                f'Entry "{entry_key}" has conflicting exchange rates: '
                f'{", ".join(conflict_details)}. '
                f'Use the same rate or fill it on only one row.'
            ),
        })
    elif len(exchange_rates_found) == 1:
        rate_str = list(exchange_rates_found.keys())[0]
        rate_row = list(exchange_rates_found.values())[0]
        try:
            exchange_rate = Decimal(rate_str)
            if exchange_rate <= 0:
                errors.append({
                    'row': rate_row,
                    'column': 'Exchange Rate',
                    'message': f'Exchange rate must be greater than zero. Found: {rate_str}.',
                })
        except (InvalidOperation, ValueError):
            errors.append({
                'row': rate_row,
                'column': 'Exchange Rate',
                'message': f'Invalid exchange rate: "{rate_str}". Must be a positive number.',
            })

    # Force rate to 1.0 for base currency
    if currency == company.base_currency:
        exchange_rate = Decimal('1.000000')

    # ── Parse each line ──
    parsed_lines = []
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')

    for row_num, row in row_list:
        # Account Name (required on every row)
        acct_name = row[COL_ACCOUNT_NAME]
        if not acct_name:
            errors.append({
                'row': row_num,
                'column': 'Account Name',
                'message': 'Account Name is required on every line.',
            })
            continue

        # Look up account
        account = account_map.get(acct_name)
        if not account:
            if acct_name in inactive_names:
                errors.append({
                    'row': row_num,
                    'column': 'Account Name',
                    'message': (
                        f'Account "{acct_name}" is inactive. '
                        f'Reactivate it first or use a different account.'
                    ),
                })
            else:
                errors.append({
                    'row': row_num,
                    'column': 'Account Name',
                    'message': (
                        f'Account "{acct_name}" not found. '
                        f'Check spelling (case-sensitive) or see the Account Reference sheet.'
                    ),
                })
            continue

        # ── Debit / Credit parsing ──
        debit_str = row[COL_DEBIT]
        credit_str = row[COL_CREDIT]

        debit_status, debit_val = _parse_amount(debit_str)
        credit_status, credit_val = _parse_amount(credit_str)

        # Invalid (non-numeric) values
        if debit_status == 'invalid':
            errors.append({
                'row': row_num, 'column': 'Debit',
                'message': f'Invalid debit amount: "{debit_str}". Must be a number.',
            })
            continue
        if credit_status == 'invalid':
            errors.append({
                'row': row_num, 'column': 'Credit',
                'message': f'Invalid credit amount: "{credit_str}". Must be a number.',
            })
            continue

        # Negative amounts
        if debit_status == 'negative':
            errors.append({
                'row': row_num, 'column': 'Debit',
                'message': (
                    f'Debit amount ({debit_val}) is negative. '
                    f'Amounts must be positive. Enter the absolute value.'
                ),
            })
            continue
        if credit_status == 'negative':
            errors.append({
                'row': row_num, 'column': 'Credit',
                'message': (
                    f'Credit amount ({credit_val}) is negative. '
                    f'Amounts must be positive. Enter the absolute value.'
                ),
            })
            continue

        # Overflow
        if debit_status == 'overflow':
            errors.append({
                'row': row_num, 'column': 'Debit',
                'message': 'Debit amount cannot exceed 999,999,999,999.99.',
            })
            continue
        if credit_status == 'overflow':
            errors.append({
                'row': row_num, 'column': 'Credit',
                'message': 'Credit amount cannot exceed 999,999,999,999.99.',
            })
            continue

        # Determine which side has a valid positive value
        # 'empty' and 'zero' both mean "this side has no amount"
        has_debit = debit_status == 'valid'
        has_credit = credit_status == 'valid'

        # Both sides zero
        if debit_status == 'zero' and credit_status == 'zero':
            errors.append({
                'row': row_num, 'column': 'Debit / Credit',
                'message': (
                    'Both Debit and Credit are zero. '
                    'Enter a positive amount greater than zero in one column.'
                ),
            })
            continue

        # Both sides have positive values
        if has_debit and has_credit:
            errors.append({
                'row': row_num, 'column': 'Debit / Credit',
                'message': (
                    f'Row has both Debit ({debit_val}) and Credit ({credit_val}). '
                    f'Each row must have only one — either Debit or Credit, not both.'
                ),
            })
            continue

        # Neither side has a value
        if not has_debit and not has_credit:
            if debit_status == 'zero' or credit_status == 'zero':
                errors.append({
                    'row': row_num, 'column': 'Debit / Credit',
                    'message': (
                        'Amount is zero. Enter a positive amount greater than zero '
                        'in either the Debit or Credit column.'
                    ),
                })
            else:
                errors.append({
                    'row': row_num, 'column': 'Debit / Credit',
                    'message': 'Row has no Debit or Credit amount. Each row must have exactly one.',
                })
            continue

        # Valid — exactly one side has a positive value
        if has_debit:
            entry_type = 'DEBIT'
            amount = debit_val
            total_debit += (amount * exchange_rate).quantize(Decimal('0.01'))
        else:
            entry_type = 'CREDIT'
            amount = credit_val
            total_credit += (amount * exchange_rate).quantize(Decimal('0.01'))

        line_desc = row[COL_LINE_DESC]

        parsed_lines.append({
            'account': account,
            'entry_type': entry_type,
            'amount': amount,
            'description': line_desc,
        })

    # ── Balance check ──
    if not errors and total_debit != total_credit:
        diff = abs(total_debit - total_credit)
        errors.append({
            'row': first_row_num,
            'column': 'Balance',
            'message': (
                f'Entry "{entry_key}" is not balanced. '
                f'Total debits: {total_debit}, Total credits: {total_credit}. '
                f'Difference: {diff}.'
            ),
        })

    parsed = None
    if not errors:
        parsed = {
            'entry_number': entry_key,
            'date': parsed_date,
            'description': description,
            'reference': reference,
            'journal_type': journal_type,
            'currency': currency,
            'exchange_rate': exchange_rate,
            'lines': parsed_lines,
        }

    return {'errors': errors, 'parsed': parsed}


# ══════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ══════════════════════════════════════════════════

def _parse_date(date_str):
    """
    Parse a date string or datetime object. Supports multiple formats.
    Returns a datetime.date object.
    Raises ValueError if unparseable.
    """
    # Handle datetime objects directly (Excel often passes these)
    import datetime as dt_module
    if isinstance(date_str, dt_module.datetime):
        return date_str.date()
    if isinstance(date_str, dt_module.date):
        return date_str

    date_str = str(date_str).strip()

    # Strip time portion if present (e.g., "2026-03-10 00:00:00")
    if ' ' in date_str:
        date_part = date_str.split(' ')[0]
        # Check if the remainder looks like a time (00:00:00)
        time_part = date_str.split(' ', 1)[1].strip()
        if ':' in time_part:
            date_str = date_part

    # Handle Excel numeric dates (float/int)
    try:
        numeric = float(date_str)
        if 40000 < numeric < 60000:
            return (dt_module.datetime(1899, 12, 30) + dt_module.timedelta(days=int(numeric))).date()
    except (ValueError, TypeError):
        pass

    # Try multiple date string formats
    formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    raise ValueError(f'Cannot parse date: {date_str}')


def _parse_amount(value_str):
    """
    Parse an amount string to Decimal with detailed status.

    Returns a tuple: (status, value)
        ('empty', None)       — blank, None, or whitespace
        ('zero', Decimal)     — explicitly zero
        ('negative', Decimal) — negative number
        ('overflow', Decimal) — exceeds max limit
        ('invalid', str)      — non-numeric garbage
        ('valid', Decimal)    — positive valid amount
    """
    if value_str is None:
        return ('empty', None)

    value_str = str(value_str).strip()
    if not value_str:
        return ('empty', None)

    # Remove commas (e.g., 1,000.00 → 1000.00)
    cleaned = value_str.replace(',', '')

    try:
        val = Decimal(cleaned).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return ('invalid', value_str)

    if val == 0:
        return ('zero', val)
    if val < 0:
        return ('negative', val)
    if val > Decimal('999999999999.99'):
        return ('overflow', val)

    return ('valid', val)