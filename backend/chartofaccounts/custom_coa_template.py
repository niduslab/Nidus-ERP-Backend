# backend/chartofaccounts/custom_coa_template.py

"""
Generates the downloadable Excel template for Custom Chart of Accounts upload.

HOW IT WORKS:
    When a user wants to use their own CoA instead of the default, they:
    1. Download this template (GET /api/companies/custom-coa-template/download/)
    2. Fill in their custom accounts following the rules
    3. Upload it during company creation (POST /api/companies/)

THE TEMPLATE HAS 4 SHEETS:
    Sheet 1 — "Instructions"
        Rules, explanations, and the list of 43 mandatory system accounts.
        Users must read this before filling in the template.

    Sheet 2 — "Default CoA Tree"
        Complete visual tree of the default 104-account CoA with all attributes.
        Helps users decide whether to use DEFAULT or CUSTOM.

    Sheet 3 — "Classifications"
        All default Layer 3 groups (marked SYSTEM — cannot remove).
        Empty rows at the bottom for users to add custom Layer 3 groups.

    Sheet 4 — "Accounts"
        All 43 system accounts pre-filled (names/codes editable, cannot remove).
        Empty rows at the bottom for users to add custom Layer 4 accounts.

CALLED FROM:
    chartofaccounts/custom_coa_views.py → CoATemplateDownloadView.get()
"""

import io

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from .seed import CLASSIFICATIONS, DEFAULT_ACCOUNTS


# ──────────────────────────────────────────────
# STYLE CONSTANTS
# ──────────────────────────────────────────────

# Fonts
FONT_TITLE = Font(name='Arial', size=16, bold=True, color='1A1A2E')
FONT_SECTION = Font(name='Arial', size=13, bold=True, color='16213E')
FONT_SUBSECTION = Font(name='Arial', size=11, bold=True, color='0F3460')
FONT_BODY = Font(name='Arial', size=10, color='333333')
FONT_BODY_BOLD = Font(name='Arial', size=10, bold=True, color='333333')
FONT_NOTE = Font(name='Arial', size=9, italic=True, color='666666')
FONT_HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_SYSTEM = Font(name='Arial', size=10, color='555555')
FONT_EDITABLE = Font(name='Arial', size=10, color='0000CC')

# Tree sheet fonts
FONT_TREE_L1 = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FONT_TREE_L2 = Font(name='Arial', size=10, bold=True, color='1A1A2E')
FONT_TREE_L3 = Font(name='Arial', size=10, bold=True, color='2C3E50')
FONT_TREE_L4 = Font(name='Arial', size=10, color='333333')
FONT_TREE_L4_SYS = Font(name='Arial', size=10, bold=True, color='0D47A1')
FONT_TREE_HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_TREE_TAG_YES = Font(name='Arial', size=9, bold=True, color='1B5E20')
FONT_TREE_TAG_NO = Font(name='Arial', size=9, color='B71C1C')
FONT_TREE_NOTE = Font(name='Arial', size=9, italic=True, color='777777')

# Fills (background colors)
FILL_HEADER = PatternFill('solid', fgColor='2C3E50')
FILL_SYSTEM_ROW = PatternFill('solid', fgColor='F0F0F0')
FILL_CUSTOM_ROW = PatternFill('solid', fgColor='E8F5E9')
FILL_SECTION_BG = PatternFill('solid', fgColor='E8EAF6')
FILL_WARNING = PatternFill('solid', fgColor='FFF3E0')

# Tree fills
FILL_TREE_L1 = PatternFill('solid', fgColor='1A237E')
FILL_TREE_L2 = PatternFill('solid', fgColor='C5CAE9')
FILL_TREE_L3 = PatternFill('solid', fgColor='E8EAF6')
FILL_TREE_L4 = PatternFill('solid', fgColor='FFFFFF')
FILL_TREE_L4_SYS = PatternFill('solid', fgColor='E3F2FD')
FILL_TREE_HEADER = PatternFill('solid', fgColor='1A237E')

# Alignments
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Border
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

TREE_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)


# ──────────────────────────────────────────────
# HELPER: Build the classification lookup
# ──────────────────────────────────────────────

def _build_classification_lookup():
    path_to_name = {}
    layer2_names = []
    # NEW: Map L3 classification name → cash_flow_category
    name_to_cash_flow = {}

    # CLASSIFICATIONS tuples now have 3 elements: (path, name, cash_flow_category)
    for internal_path, name, cash_flow_category in CLASSIFICATIONS:
        path_to_name[internal_path] = name
        if internal_path.count('.') == 1:
            layer2_names.append(name)
        if internal_path.count('.') == 2 and cash_flow_category:
            name_to_cash_flow[name] = cash_flow_category

    path_to_parent_name = {}
    for internal_path, name, _ in CLASSIFICATIONS:
        if internal_path.count('.') == 2:
            parent_path = internal_path.rsplit('.', 1)[0]
            path_to_parent_name[internal_path] = path_to_name.get(parent_path, '')

    return path_to_name, path_to_parent_name, layer2_names, name_to_cash_flow


# ──────────────────────────────────────────────
# HELPER: Apply style to a range of cells
# ──────────────────────────────────────────────

def _style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row, num_cols, is_system=False):
    fill = FILL_SYSTEM_ROW if is_system else FILL_CUSTOM_ROW
    font = FONT_SYSTEM if is_system else FONT_EDITABLE
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER


# ──────────────────────────────────────────────
# SHEET 1: INSTRUCTIONS
# ──────────────────────────────────────────────

def _build_instructions_sheet(ws):
    ws.title = "Instructions"

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 70

    row = 1

    # ── Title ──
    ws.cell(row=row, column=2, value='Nidus ERP — Custom Chart of Accounts Template').font = FONT_TITLE
    row += 2

    # ── Overview ──
    ws.cell(row=row, column=2, value='Overview').font = FONT_SECTION
    row += 1
    overview_lines = [
        'This template lets you define your own Chart of Accounts for your company.',
        'Instead of using our default CoA (with 104 pre-built accounts), you can',
        'upload only the accounts your business actually needs.',
        '',
        'Before filling this template, check the "Default CoA Tree" sheet',
        'to see the full default CoA. You may decide it already fits your needs.',
        '',
        'The CoA has a 4-layer structure:',
        '  Layer 1 — Element (Asset, Liability, Equity, Income, Expense) — FIXED',
        '  Layer 2 — Category (Current Asset, Operating Expense, etc.) — FIXED',
        '  Layer 3 — Group (Cash, Bank, Revenue, Payroll, etc.) — CUSTOMISABLE',
        '  Layer 4 — Account (Petty Cash, Sales, Bank Fees, etc.) — CUSTOMISABLE',
        '',
        'Layers 1 and 2 are fixed and cannot be changed.',
        'Layer 3 groups can be added (but default groups cannot be removed).',
        'Layer 4 is where your actual ledger accounts live.',
        '',
        'You can create custom Layer 4 accounts under BOTH default Layer 3',
        'groups AND your own custom Layer 3 groups. For example, you can:',
        '  - Add "bKash Business" under the default "Cash" group.',
        '  - Create a new group "Digital Wallets" and add accounts under it.',
        '',
        'After company creation, you can also create sub-accounts (Layer 5+)',
        'under any Layer 4 account through the application. There is no limit',
        'on sub-account depth. Sub-accounts are NOT part of this template.',
    ]
    for line in overview_lines:
        ws.cell(row=row, column=2, value=line).font = FONT_BODY
        row += 1

    row += 1

    # ── How To Fill ──
    ws.cell(row=row, column=2, value='How To Fill This Template').font = FONT_SECTION
    row += 1
    instructions = [
        'Step 1: Go to the "Classifications" sheet.',
        '   - Review the default Layer 3 groups (marked SYSTEM).',
        '   - Add any custom Layer 3 groups your business needs.',
        '   - Each custom group must specify a valid Layer 2 parent.',
        '',
        'Step 2: Go to the "Accounts" sheet.',
        '   - Review the 43 system accounts (marked SYSTEM).',
        '   - You may RENAME system accounts and change their CODES.',
        '   - You CANNOT remove any system account.',
        '   - Add your custom Layer 4 accounts in the empty rows below.',
        '   - Each account must reference a valid Layer 3 classification',
        '     (either a default group or one you added in Step 1).',
        '',
        'Step 3: Save the file and upload it during company creation.',
        '   - Choose "Custom CoA" when creating your company.',
        '   - Select this file. The system will validate it.',
        '   - If there are errors, you will get a detailed error report.',
    ]
    for line in instructions:
        ws.cell(row=row, column=2, value=line).font = FONT_BODY
        row += 1

    row += 1

    # ── Field Reference ──
    ws.cell(row=row, column=2, value='Field Reference').font = FONT_SECTION
    row += 1

    ws.cell(row=row, column=2, value='Classifications Sheet:').font = FONT_SUBSECTION
    row += 1
    class_field_lines = [
        '  Status                            — DO NOT TOUCH. Pre-filled as SYSTEM or CUSTOM.',
        '  Parent (Layer 2) (Required)       — REQUIRED for CUSTOM rows. Must exactly match a Layer 2 name below.',
        '  Classification Name (Required)    — REQUIRED. Unique name for the Layer 3 group.',
        '  Cash Flow Category (Required)     — REQUIRED. Which section of the Cash Flow Statement.',
        '                                      Options: OPERATING, INVESTING, FINANCING, CASH.',
        '                                      OPERATING = day-to-day business (most common).',
        '                                      INVESTING = long-term asset purchases/sales.',
        '                                      FINANCING = loans, equity, dividends.',
        '                                      CASH = cash and bank accounts (the result, not an activity).',
    ]
    for line in class_field_lines:
        ws.cell(row=row, column=2, value=line).font = FONT_BODY
        row += 1

    row += 1

    ws.cell(row=row, column=2, value='Accounts Sheet:').font = FONT_SUBSECTION
    row += 1
    acct_field_lines = [
        '  Status                            — DO NOT TOUCH. Pre-filled as SYSTEM or CUSTOM.',
        '  System Code                       — DO NOT TOUCH for SYSTEM rows. Leave BLANK for CUSTOM rows.',
        '  Classification (Layer 3) (Req.)   — REQUIRED. Must exactly match a name from the Classifications sheet.',
        '  Account Code (Required)           — REQUIRED. Must be unique. No spaces allowed.',
        '  Account Name (Required)           — REQUIRED. Descriptive name for the account.',
        '  Normal Balance (Required)         — REQUIRED. Exactly "DEBIT" or "CREDIT".',
        '  Currency (Optional)               — OPTIONAL. 3-letter ISO 4217 code (BDT, USD, EUR). Blank = company base currency.',
        '  Description (Optional)            — OPTIONAL. Notes about the account\'s purpose.',
    ]
    for line in acct_field_lines:
        ws.cell(row=row, column=2, value=line).font = FONT_BODY
        row += 1

    row += 1

    # ── Rules ──
    ws.cell(row=row, column=2, value='Rules').font = FONT_SECTION
    row += 1
    rules = [
        '1. All 43 system accounts MUST be present. You can rename them',
        '   and change their codes, but you cannot remove them.',
        '   They are identified by their System Code (column in Accounts sheet).',
        '',
        '2. Every account code must be UNIQUE across the entire company.',
        '   Codes cannot contain spaces.',
        '',
        '3. Every account must reference a valid Layer 3 classification.',
        '   You can use default groups OR your own custom groups.',
        '   The classification name must match exactly (case-sensitive).',
        '',
        '4. Normal Balance must be exactly "DEBIT" or "CREDIT".',
        '   For system accounts, the normal balance is pre-set and',
        '   CANNOT be changed (it is validated against expected values).',
        '',
        '5. Currency is optional. If left blank, the account will use',
        '   your company\'s base currency. If filled, it must be a valid',
        '   3-letter ISO 4217 currency code (e.g., BDT, USD, EUR, GBP).',
        '',
        '6. You cannot add new Layer 1 or Layer 2 classifications.',
        '   Only Layer 3 groups can be added.',
        '',
        '7. Default Layer 3 groups (marked SYSTEM) cannot be removed.',
        '   You can add new groups alongside them.',
        '',
        '8. Do not modify the "Status" or "System Code" columns.',
        '   These are used by the validator to identify system entries.',
    ]
    for line in rules:
        ws.cell(row=row, column=2, value=line).font = FONT_BODY
        row += 1

    row += 1

    # ── Layer 2 Parents Reference ──
    ws.cell(row=row, column=2, value='Valid Layer 2 Parents (for custom Layer 3 groups)').font = FONT_SECTION
    row += 1

    layer1_to_layer2 = {}
    for internal_path, name, _cf_cat in CLASSIFICATIONS:
        if internal_path.count('.') == 0:
            layer1_to_layer2[internal_path] = {'name': name, 'children': []}
        elif internal_path.count('.') == 1:
            parent_path = internal_path.split('.')[0]
            if parent_path in layer1_to_layer2:
                layer1_to_layer2[parent_path]['children'].append(name)

    for l1_path, l1_data in layer1_to_layer2.items():
        ws.cell(row=row, column=2, value=f'{l1_data["name"]}:').font = FONT_BODY_BOLD
        row += 1
        for child_name in l1_data['children']:
            ws.cell(row=row, column=2, value=f'    {child_name}').font = FONT_BODY
            row += 1
        row += 1

    row += 1

    # ── System Accounts Reference Table ──
    sys_count = sum(1 for a in DEFAULT_ACCOUNTS if a[6] is not None)
    ws.cell(row=row, column=2, value=f'Mandatory System Accounts ({sys_count} Total)').font = FONT_SECTION
    row += 1
    ws.cell(
        row=row, column=2,
        value='These accounts are required by ERP modules. '
              'You may rename them and change codes, but you cannot remove them.'
    ).font = FONT_NOTE
    row += 2

    path_to_name, _, _, _ = _build_classification_lookup()

    sys_headers = ['#', 'System Code', 'Default Name', 'Classification', 'Normal Balance']
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 34
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 18

    for col_idx, header in enumerate(sys_headers, start=2):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    row += 1

    sys_counter = 0
    for (
        classification_path, account_code, name,
        normal_balance, is_system, is_deletable, system_code
    ) in DEFAULT_ACCOUNTS:
        if system_code is None:
            continue

        sys_counter += 1
        classification_name = path_to_name.get(classification_path, '')

        values = [sys_counter, system_code, name, classification_name, normal_balance]
        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = THIN_BORDER
            if col_idx == 2:
                cell.alignment = ALIGN_CENTER

        row += 1

    row += 2

    ws.cell(
        row=row, column=2,
        value='Note: After company creation, you can reassign system account '
              'mappings from Settings. For example, if you want the Sales module '
              'to post to a different account, you can change it there.'
    ).font = FONT_NOTE
    row += 2

    ws.cell(
        row=row, column=2,
        value='Note: Sub-accounts (Layer 5+) can be created after company setup '
              'from the Chart of Accounts page. There is no limit on sub-account depth.'
    ).font = FONT_NOTE


# ──────────────────────────────────────────────
# SHEET 2: DEFAULT COA TREE
# ──────────────────────────────────────────────

def _build_default_coa_tree_sheet(ws):
    """
    Builds Sheet 2 — a complete visual tree of the default 104-account CoA.

    This is a reference sheet. It shows users what the DEFAULT option gives them
    so they can decide whether to use it as-is or customise.

    It is NOT parsed during upload. Purely informational.
    """
    ws.title = "Default CoA Tree"

    # Column widths
    ws.column_dimensions['A'].width = 16    # Layer 1
    ws.column_dimensions['B'].width = 26    # Layer 2
    ws.column_dimensions['C'].width = 30    # Layer 3
    ws.column_dimensions['D'].width = 10    # Account Code
    ws.column_dimensions['E'].width = 36    # Account Name
    ws.column_dimensions['F'].width = 14    # Normal Balance
    ws.column_dimensions['G'].width = 12    # Is System
    ws.column_dimensions['H'].width = 26    # System Code
    ws.column_dimensions['I'].width = 12    # Is Deletable
    ws.column_dimensions['J'].width = 52    # Reason / Notes

    # ── Header rows ──
    headers = [
        'Layer 1\n(Element)',
        'Layer 2\n(Category)',
        'Layer 3\n(Group)',
        'Account\nCode',
        'Account Name\n(Layer 4)',
        'Normal\nBalance',
        'System\nAccount?',
        'System Code\n(ERP Reference)',
        'Deletable?',
        'Reason / Notes',
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_TREE_HEADER
        cell.fill = FILL_TREE_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = TREE_BORDER

    ws.row_dimensions[1].height = 36

    # ── Build tree data ──
    path_to_name = {}
    for p, n, _cf in CLASSIFICATIONS:
        path_to_name[p] = n

    # Group accounts by Layer 3
    from collections import OrderedDict
    l3_accounts = OrderedDict()
    for (cp, code, name, nb, is_sys, is_del, sys_code) in DEFAULT_ACCOUNTS:
        if cp not in l3_accounts:
            l3_accounts[cp] = []
        l3_accounts[cp].append((code, name, nb, is_sys, is_del, sys_code))

    # Reason mapping for non-deletable accounts
    DELETION_REASONS = {
        # System accounts — reason is the ERP module dependency
        'PETTY_CASH': 'Core cash account used by payment module',
        'SALES_CASH': 'POS and sales module post collections here',
        'UNDEPOSITED_FUNDS': 'Payment module uses this as cheque clearing',
        'BANK': 'Bank module core account for reconciliation',
        'INVENTORY': 'Inventory module default posting account',
        'FINISHED_GOODS': 'Manufacturing module tracks finished goods here',
        'RAW_MATERIALS': 'Manufacturing module tracks raw materials here',
        'ADVANCE_INCOME_TAX': 'Tax module posts advance tax deductions here',
        'INPUT_VAT': 'Purchase module auto-posts input VAT here',
        'WHT_RECEIVABLE': 'Tax module auto-posts WHT deductions here',
        'ACCOUNTS_RECEIVABLE': 'Invoice module posts all credit sales here',
        'ACCUMULATED_DEPRECIATION': 'Depreciation module posts accumulated amounts here',
        'ACCOUNTS_PAYABLE': 'Purchase/bill module posts all credit purchases here',
        'ACCRUED_EXPENSE': 'System needs a default account for expense accruals',
        'WHT_PAYABLE': 'Tax module auto-posts WHT obligations here',
        'OUTPUT_VAT': 'Sales module auto-posts output VAT here',
        'INCOME_TAX_PAYABLE': 'Tax module auto-posts income tax provisions here',
        'UNEARNED_REVENUE': 'Advance payment module posts customer prepayments here',
        'CUSTOMER_DEPOSITS': 'Sales module posts customer advances here',
        'SUSPENSE': 'System clearing account for unresolved transactions',
        'CLEARING': 'System clearing account for bank reconciliation',
        'OPENING_BALANCE_OFFSET': 'System uses this during initial setup and migration',
        'OWNER_CAPITAL': 'Core equity account representing owner investment',
        'RETAINED_EARNINGS': 'Year-end closing module posts net income here',
        'DRAWING': 'Tracks owner withdrawals; essential for sole proprietorship/partnership',
        'SALES': 'Invoice module posts all revenue here',
        'SALES_DISCOUNT': 'Discount feature in sales module posts here',
        'SALES_RETURNS': 'Returns module posts credit notes here',
        'SERVICE_REVENUE': 'Sales module posts service invoices here',
        'GAIN_ON_DISPOSAL': 'Fixed asset module posts disposal gains here',
        'FX_GAIN': 'Multi-currency module auto-posts exchange gains here',
        'COGS': 'Inventory and sales module post cost of sold items here',
        'COST_OF_SERVICES': 'Service invoice module posts direct service costs here',
        'PURCHASE_DISCOUNT': 'Purchase module auto-posts early payment discounts here',
        'PURCHASE_RETURNS': 'Purchase module posts supplier returns here',
        'INVENTORY_ADJUSTMENT': 'Inventory module posts stock adjustments here',
        'DEPRECIATION_EXPENSE': 'Depreciation module posts periodic charges here',
        'AMORTISATION_EXPENSE': 'Amortisation module posts periodic charges here',
        'BANK_FEES': 'Bank module auto-posts transaction fees here',
        'INTEREST_EXPENSE': 'Loan module posts periodic interest charges here',
        'FX_LOSS': 'Multi-currency module auto-posts exchange losses here',
        'INCOME_TAX_EXPENSE': 'Tax module posts income tax charges here',
        'LOSS_ON_DISPOSAL': 'Fixed asset module posts disposal losses here',
    }

    # Non-system, non-deletable account reasons
    NON_SYSTEM_NON_DEL_REASONS = {
        'Allowance for Doubtful Accounts': 'Contra-asset required to pair with Bad Debt Expense',
        'Bad Debt Expense': 'Required to pair with Allowance for Doubtful Accounts',
    }

    # Deletable account reasons
    DELETABLE_REASONS = {
        'Mobile Banking': 'Not all businesses use mobile banking channels',
        'Working Process Inventory': 'Only needed if manufacturing module is enabled',
        'Supplementary Duty Receivable': 'Only needed for SD-applicable industries',
        'Security Deposits': 'Not all businesses pay security deposits',
        'Employee Advance': 'Only relevant if business gives salary advances',
        'Prepaid Rent': 'Only needed if business prepays rent',
        'Prepaid Insurance': 'Only needed if business prepays insurance premiums',
        'Other Advance Payments': 'General catch-all; may not be needed',
        'Notes Receivable': 'Only relevant for businesses dealing with promissory notes',
        'Loan to Others': 'Only relevant if business lends money to third parties',
        'Other Receivables': 'General catch-all; may not be needed',
        'Other Current Asset': 'General catch-all for uncategorized current assets',
        'Furnitures & Fixtures': 'Not all businesses own furniture',
        'Computers & IT Equipment': 'Not all businesses own IT equipment separately',
        'Machinery & Equipment': 'Only relevant for manufacturing or production businesses',
        'Vehicles': 'Not all businesses own vehicles',
        'Land & Buildings': 'Not all businesses own real estate',
        'Patents & Trademarks': 'Only relevant if business holds IP rights',
        'Software & License': 'Only relevant if business capitalizes software costs',
        'Goodwill': 'Only arises from business acquisitions',
        'Investments': 'Only relevant if business holds investment assets',
        'Other Non-Current Assets': 'General catch-all for uncategorized non-current assets',
        'Accrued Salaries & Wages': 'Only needed if payroll module is used',
        'Provident Fund Payable': 'Only needed if business contributes to provident fund',
        'Gratuity Payable': 'Only needed if business provides gratuity benefits',
        'Supplementary Duty Payable': 'Only needed for SD-applicable industries',
        'Other Tax & Duties Payable': 'General catch-all; may not be needed',
        'Short-Term Bank Loan': 'Not all businesses carry short-term bank loans',
        'Bank Overdraft': 'Not all businesses use overdraft facilities',
        'Credit Card Payable': 'Not all businesses use corporate credit cards',
        'Other Current Liabilities': 'General catch-all for uncategorized current liabilities',
        'General Provisions': 'Only needed if business records estimated liabilities (IAS 37)',
        'Long-Term Loans': 'Not all businesses carry long-term debt',
        'Other Long-Term Liabilities': 'General catch-all for uncategorized long-term liabilities',
        'Dividends': 'Only relevant for corporations distributing profits',
        'Shipping Charge': 'Not all businesses charge shipping separately',
        'Adjustment & Other Charges': 'Not all businesses apply adjustment charges',
        'Interest Income': 'Only relevant if business earns interest on deposits',
        'Dividend Income': 'Only relevant if business holds equity investments',
        'Rent Income': 'Only relevant if business earns rental income',
        'Commission Income': 'Only relevant if business earns agent commissions',
        'Other Income': 'General catch-all for uncategorized income',
        'Salaries and Wages': 'Not all businesses have employees',
        'Employee Benefits Expense': 'Only relevant if business provides benefits',
        'Provident Fund Expense': 'Only relevant if business contributes to provident fund',
        'Utilities Expense': 'Not all businesses pay utilities separately',
        'Rent Expense': 'Not all businesses rent premises',
        'Repair & Maintenance': 'Not all businesses incur regular maintenance costs',
        'Telephone and Internet Expense': 'Not all businesses track telecom costs separately',
        'Travel and Transportation Expense': 'Not all businesses incur travel costs',
        'Insurance Expense': 'Not all businesses carry insurance policies',
        'Office Supplies': 'Not all businesses purchase office supplies regularly',
        'Legal & Professional Fees': 'Not all businesses incur regular legal fees',
        'Rounding Adjustment': 'For payment rounding and FX cent differences',
        'Advertising & Marketing': 'Not all businesses run paid marketing campaigns',
        'Sales Commission Expense': 'Only relevant if business pays sales commissions',
        'Other Operating Expense': 'General catch-all for uncategorized operating expenses',
        'Research & Development Expense': 'Only relevant for businesses investing in R&D',
        'Other Expense': 'General catch-all for uncategorized non-operating expenses',
    }

    def _get_reason(name, sys_code, is_del):
        if sys_code and sys_code in DELETION_REASONS:
            return DELETION_REASONS[sys_code]
        if not is_del and name in NON_SYSTEM_NON_DEL_REASONS:
            return NON_SYSTEM_NON_DEL_REASONS[name]
        if is_del and name in DELETABLE_REASONS:
            return DELETABLE_REASONS[name]
        return ''

    # ── Write data ──
    row = 2
    prev_l1 = ''
    prev_l2 = ''
    prev_l3_path = ''

    # Walk through classifications in order
    for cls_path, cls_name, _cf in CLASSIFICATIONS:
        depth = cls_path.count('.')

        if depth == 0:
            # Layer 1 element
            prev_l1 = cls_name
            continue
        elif depth == 1:
            # Layer 2 category
            prev_l2 = cls_name
            continue
        elif depth == 2:
            # Layer 3 group — write accounts under it
            l3_name = cls_name
            accounts = l3_accounts.get(cls_path, [])

            if not accounts:
                # Empty Layer 3 group — still show it
                cell_l1 = ws.cell(row=row, column=1, value=prev_l1)
                cell_l2 = ws.cell(row=row, column=2, value=prev_l2)
                cell_l3 = ws.cell(row=row, column=3, value=l3_name)

                cell_l1.font = FONT_TREE_L1
                cell_l1.fill = FILL_TREE_L1
                cell_l2.font = FONT_TREE_L2
                cell_l2.fill = FILL_TREE_L2
                cell_l3.font = FONT_TREE_L3
                cell_l3.fill = FILL_TREE_L3

                for c in range(1, 11):
                    ws.cell(row=row, column=c).border = TREE_BORDER

                row += 1
                prev_l1 = ''
                prev_l2 = ''
                continue

            for i, (code, name, nb, is_sys, is_del, sys_code) in enumerate(accounts):
                # Layer 1/2/3 only show on first account in each group
                show_l1 = prev_l1
                show_l2 = prev_l2
                show_l3 = l3_name if i == 0 else ''

                c1 = ws.cell(row=row, column=1, value=show_l1 if show_l1 else None)
                c2 = ws.cell(row=row, column=2, value=show_l2 if show_l2 else None)
                c3 = ws.cell(row=row, column=3, value=show_l3 if show_l3 else None)

                ws.cell(row=row, column=4, value=code)
                ws.cell(row=row, column=5, value=name)
                ws.cell(row=row, column=6, value=nb)
                ws.cell(row=row, column=7, value='Yes' if is_sys else 'No')
                ws.cell(row=row, column=8, value=sys_code if sys_code else '—')
                ws.cell(row=row, column=9, value='Yes' if is_del else 'No')
                ws.cell(row=row, column=10, value=_get_reason(name, sys_code, is_del))

                # Styling
                if show_l1:
                    c1.font = FONT_TREE_L1
                    c1.fill = FILL_TREE_L1
                if show_l2:
                    c2.font = FONT_TREE_L2
                    c2.fill = FILL_TREE_L2
                if show_l3:
                    c3.font = FONT_TREE_L3
                    c3.fill = FILL_TREE_L3

                # Account row styling
                acct_fill = FILL_TREE_L4_SYS if is_sys else FILL_TREE_L4
                acct_font = FONT_TREE_L4_SYS if is_sys else FONT_TREE_L4

                for c in range(4, 11):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = acct_fill
                    if c in (4, 6, 7, 9):
                        cell.alignment = ALIGN_CENTER
                    else:
                        cell.alignment = ALIGN_LEFT
                    if c == 5:
                        cell.font = acct_font
                    elif c == 7:
                        cell.font = FONT_TREE_TAG_YES if is_sys else FONT_TREE_TAG_NO
                    elif c == 9:
                        cell.font = FONT_TREE_TAG_YES if is_del else FONT_TREE_TAG_NO
                    elif c == 10:
                        cell.font = FONT_TREE_NOTE

                for c in range(1, 11):
                    ws.cell(row=row, column=c).border = TREE_BORDER

                # Clear layer labels after first account
                if show_l1:
                    prev_l1 = ''
                if show_l2:
                    prev_l2 = ''

                row += 1

    # ── Summary row ──
    row += 1
    total = len(DEFAULT_ACCOUNTS)
    sys_total = sum(1 for a in DEFAULT_ACCOUNTS if a[4])
    ws.cell(row=row, column=3, value='TOTALS:').font = FONT_BODY_BOLD
    ws.cell(row=row, column=5, value=f'{total} Accounts').font = FONT_BODY_BOLD
    ws.cell(row=row, column=7, value=f'{sys_total} System').font = FONT_BODY_BOLD
    del_total = sum(1 for a in DEFAULT_ACCOUNTS if a[5])
    ws.cell(row=row, column=9, value=f'{del_total} Deletable').font = FONT_BODY_BOLD

    row += 2
    ws.cell(
        row=row, column=1,
        value='This is the complete default CoA. If you choose "Default" during '
              'company creation, you get all of these accounts. '
              'If you choose "Custom", you must keep the 43 system accounts '
              'and can add/remove the rest.'
    ).font = FONT_NOTE

    ws.freeze_panes = 'A2'


# ──────────────────────────────────────────────
# SHEET 3: CLASSIFICATIONS
# ──────────────────────────────────────────────

def _build_classifications_sheet(ws):
    ws.title = "Classifications"

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 28     # NEW: Cash Flow Category column

    headers = [
        'Status',
        'Parent (Layer 2) (Required)',
        'Classification Name (Required)',
        'Cash Flow Category (Required)',    # NEW
    ]

    _style_header_row(ws, 1, len(headers))
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    path_to_name, path_to_parent_name, _, name_to_cash_flow = _build_classification_lookup()

    row = 2
    for internal_path, name, _cf_cat in CLASSIFICATIONS:
        if internal_path.count('.') != 2:
            continue

        parent_name = path_to_parent_name.get(internal_path, '')
        # Get the cash flow category for this L3 classification
        cf_category = name_to_cash_flow.get(name, 'OPERATING')

        ws.cell(row=row, column=1, value='SYSTEM')
        ws.cell(row=row, column=2, value=parent_name)
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=cf_category)   # NEW

        _style_data_row(ws, row, len(headers), is_system=True)
        # Cash Flow Category is editable for SYSTEM rows too
        ws.cell(row=row, column=4).font = FONT_EDITABLE
        row += 1

    for _ in range(20):
        ws.cell(row=row, column=1, value='CUSTOM')
        ws.cell(row=row, column=4, value='OPERATING')    # NEW: Default for custom rows
        _style_data_row(ws, row, len(headers), is_system=False)
        row += 1


# ──────────────────────────────────────────────
# SHEET 4: ACCOUNTS
# ──────────────────────────────────────────────

def _build_accounts_sheet(ws):
    ws.title = "Accounts"

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 40

    headers = [
        'Status',
        'System Code',
        'Classification (Layer 3) (Required)',
        'Account Code (Required)',
        'Account Name (Required)',
        'Normal Balance (Required)',
        'Currency (Optional)',
        'Description (Optional)',
    ]

    _style_header_row(ws, 1, len(headers))
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    path_to_name, _, _, _ = _build_classification_lookup()

    row = 2
    for (
        classification_path, account_code, name,
        normal_balance, is_system, is_deletable, system_code
    ) in DEFAULT_ACCOUNTS:
        if system_code is None:
            continue

        classification_name = path_to_name.get(classification_path, '')

        ws.cell(row=row, column=1, value='SYSTEM')
        ws.cell(row=row, column=2, value=system_code)
        ws.cell(row=row, column=3, value=classification_name)
        ws.cell(row=row, column=4, value=account_code)
        ws.cell(row=row, column=5, value=name)
        ws.cell(row=row, column=6, value=normal_balance)
        ws.cell(row=row, column=7, value='')
        ws.cell(row=row, column=8, value='')

        _style_data_row(ws, row, len(headers), is_system=True)

        ws.cell(row=row, column=4).font = FONT_EDITABLE
        ws.cell(row=row, column=5).font = FONT_EDITABLE

        row += 1

    for _ in range(50):
        ws.cell(row=row, column=1, value='CUSTOM')
        _style_data_row(ws, row, len(headers), is_system=False)
        row += 1


# ──────────────────────────────────────────────
# MAIN: Generate the template workbook
# ──────────────────────────────────────────────

def generate_coa_template():
    """
    Creates the complete CoA template workbook and returns it as bytes.

    Returns:
        bytes: The .xlsx file content, ready to be sent as an HTTP response.
    """
    wb = Workbook()

    # Sheet 1 — Instructions
    ws_instructions = wb.active
    _build_instructions_sheet(ws_instructions)

    # Sheet 2 — Default CoA Tree (reference only)
    ws_tree = wb.create_sheet()
    _build_default_coa_tree_sheet(ws_tree)

    # Sheet 3 — Classifications
    ws_classifications = wb.create_sheet()
    _build_classifications_sheet(ws_classifications)

    # Sheet 4 — Accounts
    ws_accounts = wb.create_sheet()
    _build_accounts_sheet(ws_accounts)

    ws_classifications.freeze_panes = 'A2'
    ws_accounts.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()