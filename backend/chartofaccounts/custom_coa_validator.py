# backend/chartofaccounts/custom_coa_validator.py

"""
Validates an uploaded Custom CoA Excel file.

THE VALIDATION PIPELINE (in order):
    Phase 1 — File Structure
        Is it a valid .xlsx file?
        Does it have the required sheets?
        Do the sheets have the correct headers?

    Phase 2 — Classifications (Sheet 2)
        Are all SYSTEM classifications still present?
        Are custom classifications properly filled?
        Are parent Layer 2 references valid?
        Are there duplicates?
        Are there stray/incomplete rows?

    Phase 3 — Accounts (Sheet 3)
        Are all 43 system accounts present?
        Are account codes unique?
        Are classifications valid references?
        Are normal balances correct?
        Are currencies valid?
        Are required fields filled?
        Are there stray/incomplete rows?

CALLED FROM:
    companies/serializers.py -> CompanyCreateSerializer.validate()
"""

import openpyxl

from .seed import CLASSIFICATIONS, DEFAULT_ACCOUNTS


# ──────────────────────────────────────────────
# REFERENCE DATA
# ──────────────────────────────────────────────

CLASSIFICATION_HEADERS = ['Status', 'Parent (Layer 2) (Required)', 'Classification Name (Required)']
ACCOUNT_HEADERS = [
    'Status', 'System Code', 'Classification (Layer 3) (Required)',
    'Account Code (Required)', 'Account Name (Required)', 'Normal Balance (Required)',
    'Currency (Optional)', 'Description (Optional)',
]

# Column name mappings for user-facing error messages.
# These strip the "(Required)"/"(Optional)" suffixes to keep errors clean.
CLASSIFICATION_COL_NAMES = ['Status', 'Parent (Layer 2)', 'Classification Name']
ACCOUNT_COL_NAMES = [
    'Status', 'System Code', 'Classification (Layer 3)',
    'Account Code', 'Account Name', 'Normal Balance',
    'Currency', 'Description',
]

DEFAULT_LAYER3_NAMES = set()
DEFAULT_LAYER3_TO_PARENT = {}

_PATH_TO_NAME = {}
for _path, _name in CLASSIFICATIONS:
    _PATH_TO_NAME[_path] = _name

for _path, _name in CLASSIFICATIONS:
    if _path.count('.') == 2:
        DEFAULT_LAYER3_NAMES.add(_name)
        _parent_path = _path.rsplit('.', 1)[0]
        DEFAULT_LAYER3_TO_PARENT[_name] = _PATH_TO_NAME.get(_parent_path, '')

VALID_LAYER2_NAMES = set()
for _path, _name in CLASSIFICATIONS:
    if _path.count('.') == 1:
        VALID_LAYER2_NAMES.add(_name)

SYSTEM_ACCOUNT_REFERENCE = {}
for (
    _class_path, _code, _name, _normal_balance,
    _is_system, _is_deletable, _system_code
) in DEFAULT_ACCOUNTS:
    if _system_code is not None:
        SYSTEM_ACCOUNT_REFERENCE[_system_code] = {
            'default_name': _name,
            'classification': _PATH_TO_NAME.get(_class_path, ''),
            'normal_balance': _normal_balance,
        }

VALID_CURRENCIES = {
    'AED', 'AFN', 'ALL', 'AMD', 'ANG', 'AOA', 'ARS', 'AUD', 'AWG', 'AZN',
    'BAM', 'BBD', 'BDT', 'BGN', 'BHD', 'BIF', 'BMD', 'BND', 'BOB', 'BRL',
    'BSD', 'BTN', 'BWP', 'BYN', 'BZD', 'CAD', 'CDF', 'CHF', 'CLP', 'CNY',
    'COP', 'CRC', 'CUP', 'CVE', 'CZK', 'DJF', 'DKK', 'DOP', 'DZD', 'EGP',
    'ERN', 'ETB', 'EUR', 'FJD', 'FKP', 'GBP', 'GEL', 'GHS', 'GIP', 'GMD',
    'GNF', 'GTQ', 'GYD', 'HKD', 'HNL', 'HRK', 'HTG', 'HUF', 'IDR', 'ILS',
    'INR', 'IQD', 'IRR', 'ISK', 'JMD', 'JOD', 'JPY', 'KES', 'KGS', 'KHR',
    'KMF', 'KPW', 'KRW', 'KWD', 'KYD', 'KZT', 'LAK', 'LBP', 'LKR', 'LRD',
    'LSL', 'LYD', 'MAD', 'MDL', 'MGA', 'MKD', 'MMK', 'MNT', 'MOP', 'MRU',
    'MUR', 'MVR', 'MWK', 'MXN', 'MYR', 'MZN', 'NAD', 'NGN', 'NIO', 'NOK',
    'NPR', 'NZD', 'OMR', 'PAB', 'PEN', 'PGK', 'PHP', 'PKR', 'PLN', 'PYG',
    'QAR', 'RON', 'RSD', 'RUB', 'RWF', 'SAR', 'SBD', 'SCR', 'SDG', 'SEK',
    'SGD', 'SHP', 'SLE', 'SOS', 'SRD', 'SSP', 'STN', 'SYP', 'SZL', 'THB',
    'TJS', 'TMT', 'TND', 'TOP', 'TRY', 'TTD', 'TWD', 'TZS', 'UAH', 'UGX',
    'USD', 'UYU', 'UZS', 'VES', 'VND', 'VUV', 'WST', 'XAF', 'XCD', 'XOF',
    'XPF', 'YER', 'ZAR', 'ZMW', 'ZWL',
}


# ──────────────────────────────────────────────
# ERROR BUILDER
# ──────────────────────────────────────────────

class ValidationResult:
    def __init__(self):
        self.errors = []
        self.classifications = []
        self.accounts = []

    def add_error(self, sheet, row, column, message):
        self.errors.append({
            'sheet': sheet,
            'row': row,
            'column': column,
            'message': message,
        })

    def add_file_error(self, message):
        self.errors.append({
            'sheet': None, 'row': None, 'column': None,
            'message': message,
        })

    @property
    def is_valid(self):
        return len(self.errors) == 0

    def to_dict(self):
        return {
            'valid': self.is_valid,
            'errors': self.errors,
            'error_count': len(self.errors),
            'classifications': self.classifications if self.is_valid else [],
            'accounts': self.accounts if self.is_valid else [],
        }


# ──────────────────────────────────────────────
# HELPERS: Row reading and categorisation
# ──────────────────────────────────────────────

def _read_row(ws, row, num_cols):
    """Read a row from the worksheet and return raw values as a list."""
    return [ws.cell(row=row, column=col).value for col in range(1, num_cols + 1)]


def _clean_cell(value):
    """Clean a cell value: strip whitespace, return '' for None."""
    if value is None:
        return ''
    return str(value).strip()


def _is_row_empty(raw_values):
    """Check if ALL cells in the row are None or blank strings."""
    return all(
        v is None or (isinstance(v, str) and not v.strip())
        for v in raw_values
    )


def _find_filled_columns(raw_values, col_names):
    """
    Return a list of (column_name, cell_value) tuples for non-empty cells.
    Used to generate targeted error messages for incomplete rows.
    """
    filled = []
    for i, value in enumerate(raw_values):
        if value is not None and str(value).strip():
            name = col_names[i] if i < len(col_names) else f'Column {i + 1}'
            filled.append((name, str(value).strip()))
    return filled


# ──────────────────────────────────────────────
# PHASE 1: FILE STRUCTURE VALIDATION
# ──────────────────────────────────────────────

def _validate_file_structure(wb, result):
    sheet_names = wb.sheetnames

    ws_classifications = None
    ws_accounts = None

    for name in sheet_names:
        if name.lower() == 'classifications':
            ws_classifications = wb[name]
        elif name.lower() == 'accounts':
            ws_accounts = wb[name]

    if ws_classifications is None:
        result.add_file_error(
            'Missing required sheet: "Classifications". '
            'The workbook must have a sheet named "Classifications" '
            'containing the Layer 3 classification groups.'
        )

    if ws_accounts is None:
        result.add_file_error(
            'Missing required sheet: "Accounts". '
            'The workbook must have a sheet named "Accounts" '
            'containing the Layer 4 ledger accounts.'
        )

    if not ws_classifications or not ws_accounts:
        return None, None

    # Validate headers in Classifications sheet
    actual_class_headers = []
    for col in range(1, len(CLASSIFICATION_HEADERS) + 1):
        val = ws_classifications.cell(row=1, column=col).value
        actual_class_headers.append(str(val).strip() if val else '')

    for i, expected in enumerate(CLASSIFICATION_HEADERS):
        if i >= len(actual_class_headers) or actual_class_headers[i] != expected:
            result.add_error(
                'Classifications', 1, expected,
                f'Expected column header "{expected}" in column {i+1}, '
                f'but found "{actual_class_headers[i] if i < len(actual_class_headers) else "(empty)"}". '
                f'Do not modify the header row.'
            )

    # Validate headers in Accounts sheet
    actual_acc_headers = []
    for col in range(1, len(ACCOUNT_HEADERS) + 1):
        val = ws_accounts.cell(row=1, column=col).value
        actual_acc_headers.append(str(val).strip() if val else '')

    for i, expected in enumerate(ACCOUNT_HEADERS):
        if i >= len(actual_acc_headers) or actual_acc_headers[i] != expected:
            result.add_error(
                'Accounts', 1, expected,
                f'Expected column header "{expected}" in column {i+1}, '
                f'but found "{actual_acc_headers[i] if i < len(actual_acc_headers) else "(empty)"}". '
                f'Do not modify the header row.'
            )

    if not result.is_valid:
        return None, None

    return ws_classifications, ws_accounts


# ──────────────────────────────────────────────
# PHASE 2: CLASSIFICATIONS VALIDATION
# ──────────────────────────────────────────────

def _validate_classifications(ws, result):
    all_classification_names = set()
    system_classifications_found = set()
    name_first_seen = {}

    num_cols = len(CLASSIFICATION_HEADERS)

    for row in range(2, ws.max_row + 1):
        raw = _read_row(ws, row, num_cols)

        if _is_row_empty(raw):
            continue

        status = _clean_cell(raw[0]).upper()
        parent_name = _clean_cell(raw[1])
        class_name = _clean_cell(raw[2])

        # ── INCOMPLETE ROW DETECTION ──
        if status not in ('SYSTEM', 'CUSTOM'):
            filled = _find_filled_columns(raw, CLASSIFICATION_COL_NAMES)
            filled_names = [f'"{col}" (value: "{val}")' for col, val in filled]

            if status:
                result.add_error(
                    'Classifications', row, 'Status',
                    f'Status must be "SYSTEM" or "CUSTOM", found "{status}". '
                    f'If you are adding a custom classification, change Status to "CUSTOM" '
                    f'and fill in all required fields. If this row is not needed, '
                    f'clear all cells in the row.'
                )
            else:
                result.add_error(
                    'Classifications', row, 'Row',
                    f'Incomplete row — found data in {", ".join(filled_names)} '
                    f'but the Status column is empty. '
                    f'If you are adding a custom classification, set Status to "CUSTOM" '
                    f'and fill in all required fields: Parent (Layer 2) and Classification Name. '
                    f'If this row is not needed, clear all cells in the row.'
                )
            continue

        # ── SYSTEM classification validation ──
        if status == 'SYSTEM':
            if not class_name:
                result.add_error(
                    'Classifications', row, 'Classification Name',
                    'System classification name cannot be blank. '
                    'Do not delete system classification names.'
                )
                continue

            if class_name not in DEFAULT_LAYER3_NAMES:
                result.add_error(
                    'Classifications', row, 'Classification Name',
                    f'"{class_name}" is not a recognised system classification. '
                    f'System classification names must match exactly. '
                    f'Do not modify system classification names.'
                )
            else:
                system_classifications_found.add(class_name)

                expected_parent = DEFAULT_LAYER3_TO_PARENT.get(class_name, '')
                if parent_name and parent_name != expected_parent:
                    result.add_error(
                        'Classifications', row, 'Parent (Layer 2)',
                        f'System classification "{class_name}" must stay under '
                        f'"{expected_parent}". Found "{parent_name}" instead. '
                        f'Do not move system classifications.'
                    )

                all_classification_names.add(class_name)

        # ── CUSTOM classification validation ──
        elif status == 'CUSTOM':
            # Skip entirely empty CUSTOM rows (placeholder rows)
            if not class_name and not parent_name:
                continue

            if not class_name and parent_name:
                result.add_error(
                    'Classifications', row, 'Classification Name',
                    f'Classification Name is required. You have set Parent (Layer 2) '
                    f'to "{parent_name}" but the Classification Name is empty. '
                    f'Please provide a name for this custom classification.'
                )
                continue

            if not parent_name and class_name:
                result.add_error(
                    'Classifications', row, 'Parent (Layer 2)',
                    f'Parent (Layer 2) is required for custom classification "{class_name}". '
                    f'Valid parents: {", ".join(sorted(VALID_LAYER2_NAMES))}'
                )
                continue

            if parent_name not in VALID_LAYER2_NAMES:
                result.add_error(
                    'Classifications', row, 'Parent (Layer 2)',
                    f'"{parent_name}" is not a valid Layer 2 classification. '
                    f'Valid options: {", ".join(sorted(VALID_LAYER2_NAMES))}'
                )
            else:
                all_classification_names.add(class_name)

        # ── Duplicate detection ──
        if class_name:
            if class_name in name_first_seen:
                result.add_error(
                    'Classifications', row, 'Classification Name',
                    f'Duplicate classification name "{class_name}" '
                    f'(first appears on row {name_first_seen[class_name]}).'
                )
            else:
                name_first_seen[class_name] = row

    # ── Check all system classifications are present ──
    missing_system = DEFAULT_LAYER3_NAMES - system_classifications_found
    if missing_system:
        for name in sorted(missing_system):
            result.add_error(
                'Classifications', None, 'Classification Name',
                f'Required system classification "{name}" is missing. '
                f'Do not remove system classifications from the template.'
            )

    # Store parsed classification data
    if result.is_valid:
        for row in range(2, ws.max_row + 1):
            raw = _read_row(ws, row, num_cols)
            if _is_row_empty(raw):
                continue
            status = _clean_cell(raw[0]).upper()
            parent_name = _clean_cell(raw[1])
            class_name = _clean_cell(raw[2])

            if status == 'CUSTOM' and class_name and parent_name:
                result.classifications.append({
                    'parent_layer2_name': parent_name,
                    'name': class_name,
                    'is_custom': True,
                })

    return all_classification_names


# ──────────────────────────────────────────────
# PHASE 3: ACCOUNTS VALIDATION
# ──────────────────────────────────────────────

def _validate_accounts(ws, result, valid_classifications):
    system_accounts_found = {}
    all_codes = {}
    all_names_in_class = {}

    num_cols = len(ACCOUNT_HEADERS)

    for row in range(2, ws.max_row + 1):
        raw = _read_row(ws, row, num_cols)

        if _is_row_empty(raw):
            continue

        status = _clean_cell(raw[0]).upper()
        system_code = _clean_cell(raw[1]).upper()
        classification = _clean_cell(raw[2])
        account_code = _clean_cell(raw[3])
        account_name = _clean_cell(raw[4])
        normal_balance = _clean_cell(raw[5]).upper()
        currency = _clean_cell(raw[6]).upper()
        description = _clean_cell(raw[7])

        # ── INCOMPLETE ROW DETECTION ──
        if status not in ('SYSTEM', 'CUSTOM'):
            filled = _find_filled_columns(raw, ACCOUNT_COL_NAMES)
            filled_names = [f'"{col}" (value: "{val}")' for col, val in filled]

            if status:
                result.add_error(
                    'Accounts', row, 'Status',
                    f'Status must be "SYSTEM" or "CUSTOM", found "{status}". '
                    f'If you are adding a custom account, change Status to "CUSTOM" '
                    f'and fill in all required fields. If this row is not needed, '
                    f'clear all cells in the row.'
                )
            else:
                result.add_error(
                    'Accounts', row, 'Row',
                    f'Incomplete row — found data in {", ".join(filled_names)} '
                    f'but the Status column is empty. '
                    f'If you are adding a custom account, set Status to "CUSTOM" '
                    f'and fill in all required fields: Classification (Layer 3), '
                    f'Account Code, Account Name, and Normal Balance. '
                    f'If this row is not needed, clear all cells in the row.'
                )
            continue

        # ── SYSTEM account validation ──
        if status == 'SYSTEM':
            if not system_code:
                result.add_error(
                    'Accounts', row, 'System Code',
                    'System Code is required for SYSTEM accounts. '
                    'Do not delete the System Code value.'
                )
                continue

            if system_code not in SYSTEM_ACCOUNT_REFERENCE:
                result.add_error(
                    'Accounts', row, 'System Code',
                    f'"{system_code}" is not a recognised system code. '
                    f'Do not modify system code values.'
                )
                continue

            if system_code in system_accounts_found:
                result.add_error(
                    'Accounts', row, 'System Code',
                    f'Duplicate system account "{system_code}" '
                    f'(first appears on row {system_accounts_found[system_code]}). '
                    f'Each system account must appear exactly once.'
                )
            else:
                system_accounts_found[system_code] = row

            ref = SYSTEM_ACCOUNT_REFERENCE[system_code]

            if classification != ref['classification']:
                result.add_error(
                    'Accounts', row, 'Classification (Layer 3)',
                    f'System account "{system_code}" must be under '
                    f'classification "{ref["classification"]}". '
                    f'Found "{classification}" instead. '
                    f'Do not move system accounts to different classifications.'
                )

            if normal_balance != ref['normal_balance']:
                result.add_error(
                    'Accounts', row, 'Normal Balance',
                    f'System account "{system_code}" must have normal balance '
                    f'"{ref["normal_balance"]}". Found "{normal_balance}" instead. '
                    f'Do not change the normal balance of system accounts.'
                )

            if not account_code:
                result.add_error(
                    'Accounts', row, 'Account Code',
                    f'Account code is required for system account "{system_code}".'
                )

            if not account_name:
                result.add_error(
                    'Accounts', row, 'Account Name',
                    f'Account name is required for system account "{system_code}".'
                )

        # ── CUSTOM account validation ──
        elif status == 'CUSTOM':
            if (not system_code and not classification and not account_code
                    and not account_name and not normal_balance
                    and not currency and not description):
                continue

            if system_code:
                result.add_error(
                    'Accounts', row, 'System Code',
                    f'System Code must be blank for CUSTOM accounts. '
                    f'Found "{system_code}". Only SYSTEM accounts have system codes.'
                )

            # ── REQUIRED FIELDS CHECK ──
            missing_required = []
            if not classification:
                missing_required.append('Classification (Layer 3)')
            if not account_code:
                missing_required.append('Account Code')
            if not account_name:
                missing_required.append('Account Name')
            if not normal_balance:
                missing_required.append('Normal Balance')

            if missing_required:
                filled = _find_filled_columns(raw, ACCOUNT_COL_NAMES)
                filled_display = [
                    f'{col}: "{val}"' for col, val in filled
                    if col != 'Status'
                ]
                context = ''
                if filled_display:
                    context = f' You have filled: {", ".join(filled_display)}.'

                result.add_error(
                    'Accounts', row, 'Row',
                    f'Incomplete CUSTOM account — missing required '
                    f'field{"s" if len(missing_required) > 1 else ""}: '
                    f'{", ".join(missing_required)}.{context} '
                    f'Please complete all required fields or clear the row '
                    f'if it is not needed.'
                )

        # ── Common validations (both SYSTEM and CUSTOM) ──

        if account_code and ' ' in account_code:
            result.add_error(
                'Accounts', row, 'Account Code',
                f'Account code "{account_code}" contains spaces. '
                f'Codes cannot contain spaces.'
            )

        if account_code:
            if account_code in all_codes:
                result.add_error(
                    'Accounts', row, 'Account Code',
                    f'Duplicate account code "{account_code}" '
                    f'(first appears on row {all_codes[account_code]}).'
                )
            else:
                all_codes[account_code] = row

        if classification and classification not in valid_classifications:
            result.add_error(
                'Accounts', row, 'Classification (Layer 3)',
                f'"{classification}" is not a valid Layer 3 classification. '
                f'It must match a classification name from the Classifications sheet '
                f'(case-sensitive).'
            )

        if normal_balance and normal_balance not in ('DEBIT', 'CREDIT'):
            result.add_error(
                'Accounts', row, 'Normal Balance',
                f'Normal balance must be "DEBIT" or "CREDIT". '
                f'Found "{normal_balance}".'
            )

        if currency and currency not in VALID_CURRENCIES:
            result.add_error(
                'Accounts', row, 'Currency',
                f'"{currency}" is not a valid ISO 4217 currency code. '
                f'Leave blank to use the company\'s base currency, '
                f'or use a valid code (e.g., BDT, USD, EUR, GBP).'
            )

        if classification and account_name:
            key = (classification, account_name)
            if key in all_names_in_class:
                result.add_error(
                    'Accounts', row, 'Account Name',
                    f'Duplicate account name "{account_name}" under '
                    f'classification "{classification}" '
                    f'(first appears on row {all_names_in_class[key]}). '
                    f'Accounts under the same classification should have unique names.'
                )
            else:
                all_names_in_class[key] = row

    # ── Check all 43 system accounts are present ──
    missing_system = set(SYSTEM_ACCOUNT_REFERENCE.keys()) - set(system_accounts_found.keys())
    if missing_system:
        for code in sorted(missing_system):
            ref = SYSTEM_ACCOUNT_REFERENCE[code]
            result.add_error(
                'Accounts', None, 'System Code',
                f'Required system account "{code}" ({ref["default_name"]}) is missing. '
                f'Do not remove system accounts from the template.'
            )

    # ── Parse account data for later use ──
    if result.is_valid:
        for row in range(2, ws.max_row + 1):
            raw = _read_row(ws, row, num_cols)
            if _is_row_empty(raw):
                continue

            status = _clean_cell(raw[0]).upper()
            system_code = _clean_cell(raw[1]).upper()
            classification = _clean_cell(raw[2])
            account_code = _clean_cell(raw[3])
            account_name = _clean_cell(raw[4])
            normal_balance = _clean_cell(raw[5]).upper()
            currency = _clean_cell(raw[6]).upper()
            description = _clean_cell(raw[7])

            if status == 'CUSTOM' and not account_code and not account_name:
                continue

            if status in ('SYSTEM', 'CUSTOM'):
                result.accounts.append({
                    'is_system': status == 'SYSTEM',
                    'system_code': system_code if system_code else None,
                    'classification_name': classification,
                    'code': account_code,
                    'name': account_name,
                    'normal_balance': normal_balance,
                    'currency': currency if currency else None,
                    'description': description,
                })


# ──────────────────────────────────────────────
# MAIN ENTRY POINT
# ──────────────────────────────────────────────

def validate_coa_file(file_obj):
    result = ValidationResult()

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception:
        result.add_file_error(
            'The uploaded file is not a valid Excel (.xlsx) file. '
            'Please download the template and fill it in, then upload '
            'the saved .xlsx file.'
        )
        return result.to_dict()

    ws_classifications, ws_accounts = _validate_file_structure(wb, result)

    if not ws_classifications or not ws_accounts:
        return result.to_dict()

    valid_classification_names = _validate_classifications(ws_classifications, result)

    _validate_accounts(ws_accounts, result, valid_classification_names)

    return result.to_dict()