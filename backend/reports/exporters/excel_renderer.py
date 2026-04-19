# backend/reports/exporters/excel_renderer.py

"""
Excel (.xlsx) export renderer for all financial reports.

Uses openpyxl. Each report type has its own render function.
All share the same header block, styling, and number formatting.

FEATURES:
    - Company header block (company name, report title, dates, currency)
    - Frozen header row + auto-filter on data tables
    - Alternating row colours for readability
    - Currency-formatted number columns (right-aligned, comma-separated)
    - Bold subtotal/total rows with top border
    - Auto-adjusted column widths
    - Tree reports use indentation for nested classifications
"""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from .styles import (
    COLOUR_HEADER_BG, COLOUR_HEADER_FG, COLOUR_ROW_ALT, COLOUR_ROW_WHITE,
    COLOUR_SUBTOTAL_BG, COLOUR_TOTAL_BG, COLOUR_SECTION_BG, COLOUR_BORDER,
    COLOUR_PRIMARY, COLOUR_TEXT, COLOUR_TEXT_LIGHT,
    FONT_NAME, FONT_SIZE_TITLE, FONT_SIZE_SUBTITLE, FONT_SIZE_HEADER, FONT_SIZE_BODY,
    EXCEL_NUMBER_FORMAT,
    to_decimal, build_header_info, pl_account_amount, pl_l3_amount,
)


# ══════════════════════════════════════════════════
# SHARED EXCEL HELPERS
# ══════════════════════════════════════════════════

# Reusable style objects
_FONT_TITLE = Font(name=FONT_NAME, size=FONT_SIZE_TITLE, bold=True, color=COLOUR_PRIMARY)
_FONT_SUBTITLE = Font(name=FONT_NAME, size=FONT_SIZE_SUBTITLE, color=COLOUR_TEXT_LIGHT)
_FONT_HEADER = Font(name=FONT_NAME, size=FONT_SIZE_HEADER, bold=True, color=COLOUR_HEADER_FG)
_FONT_BODY = Font(name=FONT_NAME, size=FONT_SIZE_BODY, color=COLOUR_TEXT)
_FONT_BOLD = Font(name=FONT_NAME, size=FONT_SIZE_BODY, bold=True, color=COLOUR_TEXT)
_FONT_TOTAL = Font(name=FONT_NAME, size=FONT_SIZE_BODY, bold=True, color=COLOUR_PRIMARY)

_FILL_HEADER = PatternFill('solid', fgColor=COLOUR_HEADER_BG)
_FILL_ALT = PatternFill('solid', fgColor=COLOUR_ROW_ALT)
_FILL_WHITE = PatternFill('solid', fgColor=COLOUR_ROW_WHITE)
_FILL_SUBTOTAL = PatternFill('solid', fgColor=COLOUR_SUBTOTAL_BG)
_FILL_TOTAL = PatternFill('solid', fgColor=COLOUR_TOTAL_BG)
_FILL_SECTION = PatternFill('solid', fgColor=COLOUR_SECTION_BG)

_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

_BORDER_THIN = Border(
    bottom=Side(style='thin', color=COLOUR_BORDER),
)
_BORDER_TOP = Border(
    top=Side(style='thin', color=COLOUR_PRIMARY),
    bottom=Side(style='thin', color=COLOUR_PRIMARY),
)


def _write_header_block(ws, info, row=1):
    """Write the company/report header block at the top of the sheet."""
    ws.cell(row=row, column=1, value=info['company_name']).font = _FONT_TITLE
    row += 1

    ws.cell(row=row, column=1, value=info['report_title']).font = _FONT_BOLD
    row += 1

    # Date line
    if info.get('as_of_date'):
        date_text = 'As of {}'.format(info['as_of_date'])
    elif info.get('from_date') and info.get('to_date'):
        date_text = 'Period: {} to {}'.format(info['from_date'], info['to_date'])
    else:
        date_text = ''

    if date_text:
        ws.cell(row=row, column=1, value=date_text).font = _FONT_SUBTITLE
        row += 1

    # Currency + method line
    extra_parts = []
    if info.get('base_currency'):
        extra_parts.append('Currency: {}'.format(info['base_currency']))
    if info.get('method'):
        extra_parts.append('Method: {}'.format(info['method']))
    if info.get('filter_mode'):
        extra_parts.append('Filter: {}'.format(info['filter_mode']))
    if extra_parts:
        ws.cell(row=row, column=1, value=' | '.join(extra_parts)).font = _FONT_SUBTITLE
        row += 1

    row += 1  # Blank row before data
    return row


def _write_table_headers(ws, row, headers, col_widths=None):
    """Write a styled header row for a data table."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return row + 1


def _write_data_row(ws, row, values, is_alt=False, is_subtotal=False,
                     is_total=False, is_section=False, amount_cols=None):
    """Write a styled data row."""
    amount_cols = amount_cols or set()

    if is_total:
        fill = _FILL_TOTAL
        font = _FONT_TOTAL
        border = _BORDER_TOP
    elif is_subtotal:
        fill = _FILL_SUBTOTAL
        font = _FONT_BOLD
        border = _BORDER_THIN
    elif is_section:
        fill = _FILL_SECTION
        font = _FONT_BOLD
        border = None
    else:
        fill = _FILL_ALT if is_alt else _FILL_WHITE
        font = _FONT_BODY
        border = None

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = font
        cell.fill = fill
        if border:
            cell.border = border

        # Handle amounts
        if col_idx in amount_cols and value is not None:
            try:
                dec_val = Decimal(str(value)) if not isinstance(value, Decimal) else value
                cell.value = float(dec_val)
                cell.number_format = EXCEL_NUMBER_FORMAT
                cell.alignment = _ALIGN_RIGHT
            except Exception:
                cell.value = value
                cell.alignment = _ALIGN_LEFT
        else:
            cell.value = value
            cell.alignment = _ALIGN_LEFT

    return row + 1


def _to_bytes(wb):
    """Save workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════
# MAIN DISPATCH
# ══════════════════════════════════════════════════

def render_excel(report_type, report_data):
    """
    Main entry point. Dispatches to the appropriate report-specific renderer.
    Returns bytes (the .xlsx file content).
    """
    renderers = {
        'trial_balance': _render_trial_balance,
        'balance_sheet': _render_balance_sheet,
        'income_statement': _render_income_statement,
        'general_ledger': _render_general_ledger,
        'account_transactions': _render_account_transactions,
        'cash_flow': _render_cash_flow,
        'journal_entries': _render_journal_entries,
    }
    renderer = renderers.get(report_type)
    if not renderer:
        raise ValueError('No Excel renderer for report type: {}'.format(report_type))
    return renderer(report_data)


# ══════════════════════════════════════════════════
# TRIAL BALANCE
# ══════════════════════════════════════════════════

def _render_trial_balance(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Trial Balance'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    headers = ['Classification', 'Account Code', 'Account Name', 'Debit', 'Credit']
    widths = [30, 14, 35, 18, 18]
    row = _write_table_headers(ws, row, headers, widths)
    ws.freeze_panes = 'A{}'.format(row)
    amount_cols = {4, 5}
    alt = False

    # Walk the nested groups tree
    for group in data.get('groups', []):
        # L1 section header
        row = _write_data_row(ws, row, [group['name'], '', '', '', ''],
                               is_section=True)
        for l2 in group.get('children', []):
            # L2 header
            row = _write_data_row(ws, row, ['  ' + l2['name'], '', '', '', ''],
                                   is_section=True)
            for l3 in l2.get('children', []):
                # L3 header
                row = _write_data_row(ws, row,
                    ['    ' + l3['name'], '', '', l3.get('subtotal_debit'), l3.get('subtotal_credit')],
                    is_subtotal=True, amount_cols=amount_cols)
                # Accounts under L3 (recursive for sub-accounts)
                for acct in l3.get('accounts', []):
                    alt = not alt
                    row, alt = _write_account_row_tb_recursive(ws, row, acct, 0, alt, amount_cols)

    # Grand totals
    row += 1
    row = _write_data_row(ws, row,
        ['TOTAL', '', '', data.get('grand_total_debit'), data.get('grand_total_credit')],
        is_total=True, amount_cols=amount_cols)

    return _to_bytes(wb)


def _write_account_row_tb_recursive(ws, row, acct, depth, alt, amount_cols):
    """
    Recursively write an account row for Trial Balance.
    Shows own balance only (not subtotals) to prevent double-counting.
    Skips zero-balance intermediary accounts but still recurses into children.
    """
    own_dr = acct.get('own_debit_balance')
    own_cr = acct.get('own_credit_balance')
    has_own_balance = (own_dr is not None) or (own_cr is not None)

    # Only write a row if this account has its own balance
    if has_own_balance:
        indent = '  ' * depth
        row = _write_data_row(ws, row,
            ['',
             acct.get('code', ''),
             indent + acct.get('name', ''),
             own_dr, own_cr],
            is_alt=alt, amount_cols=amount_cols)

    # Always recurse into sub-accounts regardless
    for child in acct.get('children', []):
        alt = not alt
        row, alt = _write_account_row_tb_recursive(ws, row, child, depth + 1, alt, amount_cols)

    return row, alt


# ══════════════════════════════════════════════════
# BALANCE SHEET
# ══════════════════════════════════════════════════

def _render_balance_sheet(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Balance Sheet'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    # Single Amount column (Zoho-style)
    headers = ['Account', 'Code', 'Amount']
    widths = [45, 14, 20]
    row = _write_table_headers(ws, row, headers, widths)
    ws.freeze_panes = 'A{}'.format(row)
    amount_cols = {3}

    # Assets and Liabilities sections
    for section_name, section_key, total_key in [
        ('ASSETS', 'assets', 'total_assets'),
        ('LIABILITIES', 'liabilities', 'total_liabilities'),
    ]:
        row = _write_data_row(ws, row, [section_name, '', ''], is_section=True)
        row = _write_bs_section_tree(ws, row, data.get(section_key, []), amount_cols)
        row = _write_data_row(ws, row,
            ['Total {}'.format(section_name.title()), '', data.get(total_key)],
            is_subtotal=True, amount_cols=amount_cols)
        row += 1

    # Equity section — retained earnings before total
    row = _write_data_row(ws, row, ['EQUITY', '', ''], is_section=True)
    row = _write_bs_section_tree(ws, row, data.get('equity', []), amount_cols)

    re_auto = data.get('retained_earnings_auto', {})
    row = _write_data_row(ws, row,
        ['  Current Year Net Income (auto)', '', re_auto.get('current_year_earnings')],
        is_subtotal=True, amount_cols=amount_cols)
    row = _write_data_row(ws, row,
        ['  Prior Year Retained (auto)', '', re_auto.get('prior_year_retained')],
        is_subtotal=True, amount_cols=amount_cols)

    row = _write_data_row(ws, row,
        ['Total Equity', '', data.get('total_equity')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # Grand totals
    row = _write_data_row(ws, row,
        ['Total Assets', '', data.get('total_assets')],
        is_total=True, amount_cols=amount_cols)
    row = _write_data_row(ws, row,
        ['Total Liabilities & Equity', '', data.get('total_liabilities_and_equity')],
        is_total=True, amount_cols=amount_cols)

    balanced = 'YES' if data.get('is_balanced') else 'NO'
    ws.cell(row=row + 1, column=1, value='Balanced: {}'.format(balanced)).font = _FONT_BOLD

    return _to_bytes(wb)


def _write_bs_section_tree(ws, row, section, amount_cols):
    """Write L2 > L3 > account tree with single Amount column for Balance Sheet."""
    alt = False
    for l2 in section:
        row = _write_data_row(ws, row, ['  ' + l2['name'], '', ''], is_section=True)
        for l3 in l2.get('children', []):
            l3_amt = pl_l3_amount(l3)
            row = _write_data_row(ws, row,
                ['    ' + l3['name'], '', l3_amt],
                is_subtotal=True, amount_cols=amount_cols)
            for acct in l3.get('accounts', []):
                alt = not alt
                row, alt = _write_pl_account_recursive(ws, row, acct, 3, alt, amount_cols)
    return row


def _write_account_recursive(ws, row, acct, depth, alt, amount_cols):
    """
    Recursively write an account row and all its sub-account children.
    Shows own balance only (not subtotals) to prevent double-counting.
    Skips zero-balance intermediary accounts but still recurses into children.
    """
    own_dr = acct.get('own_debit_balance')
    own_cr = acct.get('own_credit_balance')
    has_own_balance = (own_dr is not None) or (own_cr is not None)

    if has_own_balance:
        indent = '  ' * depth
        row = _write_data_row(ws, row,
            [indent + acct.get('name', ''),
             acct.get('code', ''),
             own_dr, own_cr],
            is_alt=alt, amount_cols=amount_cols)

    for child in acct.get('children', []):
        alt = not alt
        row, alt = _write_account_recursive(ws, row, child, depth + 1, alt, amount_cols)

    return row, alt


# ══════════════════════════════════════════════════
# INCOME STATEMENT (Zoho Books-style P&L — single Amount column)
# ══════════════════════════════════════════════════

def _write_pl_account_recursive(ws, row, acct, depth, alt, amount_cols):
    """
    Recursively write a P&L account row with a single Amount column.
    Shows own_balance as a single positive number (sign is handled by totals).
    Skips zero-balance intermediary accounts.
    """
    amt = pl_account_amount(acct)
    if amt is not None:
        indent = '  ' * depth
        row = _write_data_row(ws, row,
            [indent + acct.get('name', ''), acct.get('code', ''), amt],
            is_alt=alt, amount_cols=amount_cols)

    for child in acct.get('children', []):
        alt = not alt
        row, alt = _write_pl_account_recursive(ws, row, child, depth + 1, alt, amount_cols)

    return row, alt


def _write_pl_section(ws, row, section_name, section_data, amount_cols):
    """
    Write one P&L section with single Amount column.
    Shows L3 groups with net amount + accounts with own balance.
    """
    row = _write_data_row(ws, row, [section_name, '', ''], is_section=True)
    alt = False
    for l2 in section_data:
        for l3 in l2.get('children', []):
            # L3 classification with net amount
            l3_amt = pl_l3_amount(l3)
            row = _write_data_row(ws, row,
                ['  ' + l3['name'], '', l3_amt],
                is_subtotal=True, amount_cols=amount_cols)
            # Accounts under L3 (recursive, single amount)
            for acct in l3.get('accounts', []):
                alt = not alt
                row, alt = _write_pl_account_recursive(
                    ws, row, acct, 2, alt, amount_cols)
    return row


def _render_income_statement(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Income Statement'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    # Single Amount column instead of Debit + Credit
    headers = ['Account', 'Code', 'Amount']
    widths = [45, 14, 20]
    row = _write_table_headers(ws, row, headers, widths)
    ws.freeze_panes = 'A{}'.format(row)
    amount_cols = {3}

    # ── Operating Income ──
    row = _write_pl_section(ws, row, 'OPERATING INCOME',
                            data.get('operating_income', []), amount_cols)
    row = _write_data_row(ws, row,
        ['Total Operating Income', '', data.get('total_operating_income')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # ── Cost of Goods Sold ──
    row = _write_pl_section(ws, row, 'COST OF GOODS SOLD',
                            data.get('cost_of_goods_sold', []), amount_cols)
    row = _write_data_row(ws, row,
        ['Total Cost of Goods Sold', '', data.get('total_cogs')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # ── Gross Profit ──
    row = _write_data_row(ws, row,
        ['GROSS PROFIT', '', data.get('gross_profit')],
        is_total=True, amount_cols=amount_cols)
    row += 1

    # ── Operating Expenses ──
    row = _write_pl_section(ws, row, 'OPERATING EXPENSES',
                            data.get('operating_expenses', []), amount_cols)
    row = _write_data_row(ws, row,
        ['Total Operating Expenses', '', data.get('total_operating_expenses')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # ── Operating Profit ──
    row = _write_data_row(ws, row,
        ['OPERATING PROFIT', '', data.get('operating_profit')],
        is_total=True, amount_cols=amount_cols)
    row += 1

    # ── Non-Operating Income ──
    row = _write_pl_section(ws, row, 'NON-OPERATING INCOME',
                            data.get('non_operating_income', []), amount_cols)
    row = _write_data_row(ws, row,
        ['Total Non-Operating Income', '', data.get('total_non_operating_income')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # ── Non-Operating Expenses ──
    row = _write_pl_section(ws, row, 'NON-OPERATING EXPENSES',
                            data.get('non_operating_expenses', []), amount_cols)
    row = _write_data_row(ws, row,
        ['Total Non-Operating Expenses', '', data.get('total_non_operating_expenses')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # ── Net Profit / Loss ──
    row = _write_data_row(ws, row,
        ['NET PROFIT / LOSS', '', data.get('net_income')],
        is_total=True, amount_cols=amount_cols)

    return _to_bytes(wb)


# ══════════════════════════════════════════════════
# GENERAL LEDGER
# ══════════════════════════════════════════════════

def _render_general_ledger(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'General Ledger'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    headers = ['Date', 'Account Code', 'Account Name', 'Debit', 'Credit',
               'Running Balance', 'Note', 'Journal Type']
    widths = [14, 14, 30, 16, 16, 18, 30, 14]
    row = _write_table_headers(ws, row, headers, widths)
    ws.freeze_panes = 'A{}'.format(row)
    amount_cols = {4, 5, 6}
    alt = False

    for account in data.get('accounts', []):
        # Account header row
        row = _write_data_row(ws, row,
            ['', account['code'], account['name'],
             '', '', account.get('opening_balance'), 'Opening Balance', ''],
            is_section=True, amount_cols=amount_cols)

        for txn in account.get('transactions', []):
            alt = not alt
            # Split entry_type + base_amount into separate Debit/Credit cells
            dr_val = txn.get('debit')
            cr_val = txn.get('credit')
            row = _write_data_row(ws, row,
                [txn.get('date'), account['code'], account['name'],
                 dr_val, cr_val,
                 txn.get('running_balance'), txn.get('note'), txn.get('journal_type')],
                is_alt=alt, amount_cols=amount_cols)

        # Closing row
        row = _write_data_row(ws, row,
            ['', account['code'], account['name'],
             account.get('total_debit'), account.get('total_credit'),
             account.get('closing_balance'), 'Closing Balance', ''],
            is_subtotal=True, amount_cols=amount_cols)
        row += 1  # Blank row between accounts

    # Grand totals
    row = _write_data_row(ws, row,
        ['GRAND TOTAL', '', '',
         data.get('grand_total_debit'), data.get('grand_total_credit'), '', '', ''],
        is_total=True, amount_cols=amount_cols)

    return _to_bytes(wb)


# ══════════════════════════════════════════════════
# ACCOUNT TRANSACTIONS
# ══════════════════════════════════════════════════

def _render_account_transactions(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Account Transactions'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    # Account info sub-header
    acct = data.get('account', {})
    ws.cell(row=row, column=1,
            value='Account: {} — {} ({})'.format(
                acct.get('code', ''), acct.get('name', ''), acct.get('normal_balance', '')
            )).font = _FONT_BOLD
    row += 2

    headers = ['Date', 'Source #', 'Description', 'Reference', 'Debit',
               'Credit', 'Running Balance', 'Type']
    widths = [14, 14, 35, 18, 16, 16, 18, 14]
    row = _write_table_headers(ws, row, headers, widths)
    ws.freeze_panes = 'A{}'.format(row)
    amount_cols = {5, 6, 7}

    # Opening balance row
    row = _write_data_row(ws, row,
        ['', '', 'Opening Balance', '', '', '', data.get('opening_balance'), ''],
        is_section=True, amount_cols=amount_cols)

    alt = False
    for txn in data.get('transactions', []):
        alt = not alt
        dr_val = txn.get('debit')
        cr_val = txn.get('credit')
        row = _write_data_row(ws, row,
            [txn.get('date'), txn.get('source_number'), txn.get('source_description'),
             txn.get('source_reference'), dr_val,
             cr_val, txn.get('running_balance'), txn.get('journal_type')],
            is_alt=alt, amount_cols=amount_cols)

    # Closing balance row
    row += 1
    row = _write_data_row(ws, row,
        ['', '', 'Closing Balance', '',
         data.get('total_debit'), data.get('total_credit'),
         data.get('closing_balance'), ''],
        is_total=True, amount_cols=amount_cols)

    return _to_bytes(wb)


# ══════════════════════════════════════════════════
# CASH FLOW STATEMENT
# ══════════════════════════════════════════════════

def _render_cash_flow(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cash Flow Statement'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    headers = ['Item', 'Amount']
    widths = [50, 20]
    row = _write_table_headers(ws, row, headers, widths)
    amount_cols = {2}

    # Operating Activities
    oa = data.get('operating_activities', {})
    row = _write_data_row(ws, row, ['OPERATING ACTIVITIES', ''], is_section=True)
    row = _write_data_row(ws, row, ['  Net Income', oa.get('net_income')], amount_cols=amount_cols)
    adj = oa.get('adjustments', {})
    row = _write_data_row(ws, row, ['  Depreciation & Amortisation (add-back)', adj.get('depreciation_and_amortization')], amount_cols=amount_cols)

    row = _write_data_row(ws, row, ['  Working Capital Changes:', ''], is_subtotal=True)
    for item in oa.get('working_capital_changes', []):
        row = _write_data_row(ws, row,
            ['    {} ({})'.format(item['name'], item['code']), item.get('cash_effect')],
            amount_cols=amount_cols)
    row = _write_data_row(ws, row,
        ['  Net Cash from Operating Activities', oa.get('net_cash_from_operating')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # Investing Activities
    ia = data.get('investing_activities', {})
    row = _write_data_row(ws, row, ['INVESTING ACTIVITIES', ''], is_section=True)
    for item in ia.get('items', []):
        row = _write_data_row(ws, row,
            ['  {} ({})'.format(item['name'], item['code']), item.get('cash_effect')],
            amount_cols=amount_cols)
    row = _write_data_row(ws, row,
        ['  Net Cash from Investing Activities', ia.get('net_cash_from_investing')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # Financing Activities
    fa = data.get('financing_activities', {})
    row = _write_data_row(ws, row, ['FINANCING ACTIVITIES', ''], is_section=True)
    for item in fa.get('items', []):
        row = _write_data_row(ws, row,
            ['  {} ({})'.format(item['name'], item['code']), item.get('cash_effect')],
            amount_cols=amount_cols)
    row = _write_data_row(ws, row,
        ['  Net Cash from Financing Activities', fa.get('net_cash_from_financing')],
        is_subtotal=True, amount_cols=amount_cols)
    row += 1

    # Cash Reconciliation
    cr = data.get('cash_reconciliation', {})
    row = _write_data_row(ws, row, ['CASH RECONCILIATION', ''], is_section=True)
    row = _write_data_row(ws, row, ['  Opening Cash Balance', cr.get('opening_cash_balance')], amount_cols=amount_cols)
    row = _write_data_row(ws, row, ['  Net Change in Cash', cr.get('net_change_in_cash')], amount_cols=amount_cols)
    row = _write_data_row(ws, row,
        ['  Closing Cash Balance', cr.get('closing_cash_balance')],
        is_total=True, amount_cols=amount_cols)

    balanced = 'YES' if cr.get('is_balanced') else 'NO'
    ws.cell(row=row + 1, column=1, value='Balanced: {}'.format(balanced)).font = _FONT_BOLD

    return _to_bytes(wb)


# ══════════════════════════════════════════════════
# JOURNAL ENTRIES
# ══════════════════════════════════════════════════

def _render_journal_entries(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Journal Entries'
    info = build_header_info(data)
    row = _write_header_block(ws, info)

    headers = ['Entry #', 'Date', 'Status', 'Type', 'Description', 'Reference',
               'Account Code', 'Account Name', 'Debit', 'Credit']
    widths = [14, 14, 10, 14, 35, 18, 14, 30, 16, 16]
    row = _write_table_headers(ws, row, headers, widths)
    ws.freeze_panes = 'A{}'.format(row)
    amount_cols = {9, 10}
    alt = False

    for journal in data.get('journals', []):
        for line in journal.get('lines', []):
            alt = not alt
            dr_val = line.get('amount') if line.get('entry_type') == 'DEBIT' else None
            cr_val = line.get('amount') if line.get('entry_type') == 'CREDIT' else None
            row = _write_data_row(ws, row,
                [journal.get('entry_number'), journal.get('date'),
                 journal.get('status'), journal.get('journal_type'),
                 journal.get('description'), journal.get('reference'),
                 line.get('account_code'), line.get('account_name'),
                 dr_val, cr_val],
                is_alt=alt, amount_cols=amount_cols)

    return _to_bytes(wb)